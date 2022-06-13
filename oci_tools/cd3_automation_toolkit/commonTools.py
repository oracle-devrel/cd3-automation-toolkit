import pandas as pd
import os
import shutil
import datetime
import configparser
import oci
from oci.identity import IdentityClient
from oci.config import DEFAULT_LOCATION
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side
from contextlib import contextmanager
import collections
import re
import json as simplejson
import warnings
warnings.simplefilter("ignore")

def data_frame(filename,sheetname):

    # Read the tab from excel, Drop null values, Reset index
    df, col_headers = commonTools.read_cd3(filename, sheetname)
    df = df.dropna(how='all')
    df = df.reset_index(drop=True)

    return df


class commonTools():
    endNames = {'<END>', '<end>', '<End>'}
    tagColumns = {'freeform tags', 'freeform_tags', 'defined_tags', 'defined tags'}
    drg_auto_RTs = {'Autogenerated Drg Route Table for RPC, VC, and IPSec attachments', 'Autogenerated Drg Route Table for VCN attachments'}
    drg_auto_RDs = {'Autogenerated Import Route Distribution for ALL routes', 'Autogenerated Import Route Distribution for VCN Routes'}

    #Read Regions and Protocols Files and Excel_Columns and create dicts
    def __init__(self):
        self.all_regions=[]
        self.home_region=""
        self.ntk_compartment_ids = {}
        self.region_dict={}
        self.protocol_dict={}
        self.sheet_dict={}

        # When called from wthin OCSWorkVM or user-scripts
        dir=os.getcwd()
        if ("OCSWorkVM" in os.getcwd() or 'user-scripts' in os.getcwd()):
            os.chdir("../")
        regionFileName="OCI_Regions"
        protocolFileName="OCI_Protocols"
        excelColumnName="Excel_Columns"
        with open (regionFileName) as f:
            for line in f:
                if("#" not in line and ":" in line):
                    key = line.split(":")[0].strip().lower()
                    val = line.split(":")[1].strip().lower()
                    self.region_dict[key] = val

        with open(protocolFileName) as f:
            for line in f:
                if ("#" not in line and ":" in line):
                    key=line.split(":")[0].strip()
                    val=line.split(":")[1].strip()
                    self.protocol_dict[key]=val

        #Get Dict for column names
        with open(excelColumnName) as f:
            s = f.read()
        while True:
            decoder = simplejson.JSONDecoder()
            obj, end = decoder.raw_decode(s)
            self.sheet_dict = dict(obj)
            break

        #Change back to Initial
        if ("OCSWorkVM" in dir):
            os.chdir(dir)
        #os.chdir(dir)

    #Get Tenancy Regions
    def get_subscribedregions(self, configFileName=DEFAULT_LOCATION):
        #Get config client
        config = oci.config.from_file(file_location=configFileName)
        idc = IdentityClient(config, retry_strategy=oci.retry.DEFAULT_RETRY_STRATEGY)
        regionsubscriptions = idc.list_region_subscriptions(tenancy_id=config['tenancy'])
        homeregion=""
        for rs in regionsubscriptions.data:
            if (rs.is_home_region == True):
                homeregion = rs.region_name
            for k, v in self.region_dict.items():
                if (homeregion!="" and v == homeregion):
                    self.home_region = k
                if (rs.region_name == v):
                    self.all_regions.append(k)


    #Get Compartment OCIDs
    def get_network_compartment_ids(self,c_id, c_name,configFileName):
        # Get config client
        if configFileName == "":
            config = oci.config.from_file()
        else:
            config = oci.config.from_file(file_location=configFileName)

        tenancy_id=config['tenancy']
        idc = IdentityClient(config,retry_strategy=oci.retry.DEFAULT_RETRY_STRATEGY)
        compartments = oci.pagination.list_call_get_all_results(idc.list_compartments,compartment_id=c_id, compartment_id_in_subtree=False)

        for c in compartments.data:
            if c.lifecycle_state == "ACTIVE":
                if (c_name != "root"):
                    name = c_name + "::" + c.name
                else:
                    name = c.name
                self.ntk_compartment_ids[name] = c.id
                # Put individual compartment names also in the dictionary
                if (name != c.name):
                    if c.name not in self.ntk_compartment_ids.keys():
                        self.ntk_compartment_ids[c.name] = c.id
                    else:
                        # Remove the individual name added to dict as it is duplicate only if its parent comp is not root
                        c_details = idc.get_compartment(self.ntk_compartment_ids[c.name]).data
                        if (c_details.compartment_id != tenancy_id):
                            self.ntk_compartment_ids.pop(c.name)

                self.get_network_compartment_ids(c.id, name,configFileName)

        self.ntk_compartment_ids["root"]=tenancy_id
        del tenancy_id
        del idc
        del config

    def get_comp_list_for_export(input_compartment_names, ntk_compartment_ids):

        remove_comps = []
        comp_list_fetch = []
        if input_compartment_names is not None:
            for x in range(0, len(input_compartment_names)):
                if (input_compartment_names[x] not in ntk_compartment_ids.keys()):
                    print("Input compartment: " + input_compartment_names[x] + " doesn't exist in OCI")
                    remove_comps.append(input_compartment_names[x])

            input_compartment_names = [x for x in input_compartment_names if x not in remove_comps]
            if (len(input_compartment_names) == 0):
                print("None of the input compartments specified exist in OCI..Exiting!!!")
                exit(1)
            else:
                print("Fetching for Compartments... " + str(input_compartment_names))
                comp_list_fetch = input_compartment_names
        else:
            print("Fetching for all Compartments...")
            comp_ocids = []
            for key, val in ntk_compartment_ids.items():
                if val not in comp_ocids:
                    comp_ocids.append(val)
                    comp_list_fetch.append(key)
        return comp_list_fetch

    #Check value exported
    #If None - replace with ""
    #If list, convert to comma sepearted string
    def check_exported_value(value):
        if value == None:
            value = ""
        if ("list" in str(type(value))):
            str1 = ""
            if(value.__len__()==0):
                value=""
            for v in value:
                str1 = v + "," + str1
            if (str1 != "" and str1[-1] == ','):
                value = str1[:-1]

        return value

    # Export Tag fields - common code - Defined and Freeform Tags
    # header - individual headers/column name
    # values_for_column - list of columns from read_cd3 function
    def export_tags(oci_obj, header, values_for_column):
        defined_tags=""
        freeform_tags=""
        if 'defined' in str(header).lower():
            if (oci_obj.__getattribute__('defined_tags')):
                for namespace, tags in oci_obj.__getattribute__('defined_tags').items():
                    for key, value in tags.items():
                        # Each Namespace/TagKey - Value pair ends with a ;
                        value = str(namespace + "." + key + "=" + value)
                        defined_tags=value+";"+defined_tags

            if (defined_tags != "" and defined_tags[-1] == ';'):
                defined_tags = defined_tags[:-1]
            values_for_column[header].append(defined_tags)

        elif 'free' in str(header).lower():
            if (oci_obj.__getattribute__('freeform_tags')):
                for keys, values in oci_obj.__getattribute__('freeform_tags').items():
                    value = str(keys + "=" + values)
                    freeform_tags= value + ';' +freeform_tags

            if (freeform_tags != '' and freeform_tags[-1] == ';'):
                freeform_tags = freeform_tags [:-1]
            values_for_column[header].append(freeform_tags)

        return values_for_column

    #Export extra fields
    def export_extra_columns(oci_objs, col_header, sheet_dict, values_for_column):
        value = ""
        if (col_header in sheet_dict.keys()):
            # Check if property exists for any object on that sheet
            for oci_obj in oci_objs:
                try:
                    value = oci_obj.__getattribute__(sheet_dict[col_header])
                    value = commonTools.check_exported_value(value)
                    break
                except AttributeError as e:
                    pass
            values_for_column[col_header].append(value)
        # For Cols not defined in Excel_Columns sheet
        else:
            # Check if property exists for any object on that sheet
            for oci_obj in oci_objs:
                try:
                    value = oci_obj.__getattribute__(commonTools.check_column_headers(col_header))
                    value = commonTools.check_exported_value(value)
                    break
                except AttributeError as e:
                    pass
            values_for_column[col_header].append(value)

        return values_for_column

    # Check CD3 Column headers
    def check_column_headers(var_name):
        # replace special characters and spaces with '_' and convert to lowercase
        # replaces multiple occurrence of '_' to just 1
        var_name = var_name.strip()
        var_name = re.sub('[@!#$%^&*<>?/}{~: \n()|]', '_', var_name).lower()
        var_name = re.sub('_+', '_', var_name).lower()
        return var_name

    #Check TF variable Name
    def check_tf_variable(var_name):
        tfname = re.compile('[^a-zA-Z0-9_-]')
        tfnamestart = re.compile('[A-Za-z]')

        var_name = tfname.sub("-", var_name)
        x = tfnamestart.match(var_name)
        # variable name doesnot start with letter; append with c
        if (x == None):
            var_name = "c" + var_name
        return var_name

    # Process ColumnValues
    def check_columnvalue(columnvalue):

        if str(columnvalue).lower() == 'true' or str(columnvalue).lower() == 'false':
            columnvalue = str(columnvalue).lower()

        if (columnvalue.lower() == 'nan'):
            columnvalue = ""

        # replace \ with \\
        if("\\" in columnvalue):
            columnvalue = columnvalue.replace("\\", "\\\\")

        # replace " with \"
        if("\"" in columnvalue):
            columnvalue=columnvalue.replace("\"","\\\"")

        return columnvalue

    # Process column values with ::
    def check_multivalues_columnvalue(columnvalue, columnname, tempdict):
        columnvalue = str(columnvalue).strip()
        columnname = commonTools.check_column_headers(columnname)
        if "::" in columnvalue:
            if ".Flex" in columnvalue:
                columnname = commonTools.check_column_headers(columnname)
                multivalues = columnvalue.split("::")
                multivalues = [str(part).strip() for part in multivalues if part]
                tempdict = {columnname: multivalues}
            elif columnname != 'Compartment Name':
                columnname = commonTools.check_column_headers(columnname)
                multivalues = columnvalue.split("::")
                multivalues = [str(part).strip() for part in multivalues ]#if part]
                tempdict = {columnname: multivalues}
        return tempdict

    # Split values for tagging
    def split_tag_values(columnname, columnvalue, tempdict):
        columnvalue = columnvalue.replace("\n", "")
        if ";" in columnvalue:
            # If there are more than one tag; split them by ";" and "="
            columnname = commonTools.check_column_headers(columnname)
            multivalues = columnvalue.split(";")
            multivalues = [part.split("=") for part in multivalues if part]
            for value in multivalues:
                try:
                    value[1] = value[1].replace("\\","\\\\")
                except IndexError as e:
                    pass
            tempdict = {columnname: multivalues}
        else:
            # If there is only one tag; split them only by "="; each key-value pair is stored as a list
            columnname = commonTools.check_column_headers(columnname)
            multivalues = columnvalue.split("=")
            multivalues = [str(part).strip() for part in multivalues if part]
            if multivalues != []:
                try:
                    multivalues[1] = multivalues[1].replace("\\","\\\\")
                except IndexError as e:
                    pass
            tempdict = {columnname: [multivalues]}
        return tempdict

    # Read rows from CD3
    def read_cd3(cd3file, sheet_name):
        df = {}
        try:
            df = pd.read_excel(cd3file, sheet_name=sheet_name, skiprows=1, dtype=object)

        except Exception as e:
            if("Events" in str(e) or "Notifications" in str(e)):
                print("\nTabs - \"Events\" or \"Notifications\" is missing in the CD3. Please make sure to use the correct input file for Events and Notifications in properties file...Exiting!!")
                exit(1)
            else:
                print("Error occurred while reading the CD3 excel sheet: "+ str(e))
                exit(1)

        yield df
        try:
            book = load_workbook(cd3file)
            sheet = book[sheet_name]
        except KeyError as e:
            if 'does not exist' in str(e):
                print("\nTab - \""+sheet_name+"\" seems to be missing in the CD3. Please make sure to use the right CD3 in properties file.....Exiting!!")
                exit(1)
        except Exception as e:
            print(str(e))
            print("Exiting!!")
            exit()

        values_for_column = collections.OrderedDict()
        # values_for_column={}
        for j in range(0, sheet.max_column):
            col_name = sheet.cell(row=2, column=j + 1).value
            if (type(col_name) == str):
                values_for_column[col_name] = []
        yield values_for_column

    #Write exported  rows to cd3
    def write_to_cd3(values_for_column, cd3file, sheet_name):
        try:
            book = load_workbook(cd3file)
            sheet = book[sheet_name]
        except Exception as e:
            print(str(e))
            print("Exiting!!")
            exit()
        if (sheet_name == "VCN Info"):
            onprem_destinations = ""
            ngw_destinations = ""
            igw_destinations = ""
            for destination in values_for_column["onprem_destinations"]:
                onprem_destinations=destination+","+onprem_destinations
            for destination in values_for_column["ngw_destinations"]:
                ngw_destinations = destination + "," + ngw_destinations
            for destination in values_for_column["igw_destinations"]:
                igw_destinations = destination + "," + igw_destinations

            if (onprem_destinations != "" and onprem_destinations[-1] == ','):
                onprem_destinations = onprem_destinations[:-1]
            if (ngw_destinations != "" and ngw_destinations[-1] == ','):
                ngw_destinations = ngw_destinations[:-1]
            if (igw_destinations != "" and igw_destinations[-1] == ','):
                igw_destinations = igw_destinations[:-1]

            sheet.cell(3,2).value = onprem_destinations
            sheet.cell(4,2).value = ngw_destinations
            sheet.cell(5,2).value = igw_destinations
            try:
                book.save(cd3file)
                book.close()
            except Exception as e:
                print(str(e))
                print("Exiting!!")
                exit()
            return


        #rows_len=len(rows)
        rows_len = len(values_for_column["Region"])

        #If no rows exported from OCI, remove the sample data as well
        if(rows_len == 0):
            print("0 rows exported; Nothing to write to CD3 excel; Tab "+sheet_name +" will be empty in CD3 excel!!")
            for i in range(0, sheet.max_row):
                for j in range(0, sheet.max_column):
                    sheet.cell(row=i + 3, column=j + 1).value = ""
            try:
                book.save(cd3file)
                book.close()
            except Exception as e:
                print(str(e))
                print("Exiting!!")
                exit()
            return

        sheet_max_rows = sheet.max_row
        if (rows_len > sheet_max_rows):
            large = rows_len
        else:
            large = sheet_max_rows

        #Put Data
        j=0
        for i in range(0,large):
            for col_name in values_for_column.keys():
                if(i>=rows_len):
                    sheet.cell(row=i+3, column=j+1).value = ""
                else:
                    sheet.cell(row=i+3, column=j+1).value = values_for_column[col_name][i]
                sheet.cell(row=i+3, column=j+1).alignment = Alignment(wrap_text=True)
                j=j+1
            j=0


        brdr = Border(left=Side(style='thin'),
                      right=Side(style='thin'),
                      top=Side(style='thin'),
                      bottom=Side(style='thin'),
                      )

        for row in sheet.iter_rows(min_row=3):
            for cell in row:
                cell.border = brdr

        # Add color for exported sec rules and route rules
        if (sheet_name == "RouteRulesinOCI" or sheet_name == "SecRulesinOCI" or sheet_name == "DRGRouteRulesinOCI"):
            names = []
            # Add color coding to exported rules
            for row in sheet.iter_rows(min_row=3):
                c = 0
                region = ""
                name = ""
                for cell in row:
                    c = c + 1
                    if (c == 1):
                        region = cell.value
                        continue
                    elif (c == 4):
                        name = cell.value
                        break

                vcn_name = region + "_" + name
                if (vcn_name not in names):
                    names.append(vcn_name)
                    for cellnew in row:
                        if (len(names) % 2 == 0):
                            cellnew.fill = PatternFill(start_color="94AFAF", end_color="94AFAF", fill_type="solid")
                            cellnew.border = brdr
                        else:
                            cellnew.fill = PatternFill(start_color="E5DBBE", end_color="E5DBBE", fill_type="solid")
                            cellnew.border = brdr
                else:
                    for cellnew in row:
                        if (len(names) % 2 == 0):
                            cellnew.fill = PatternFill(start_color="94AFAF", end_color="94AFAF", fill_type="solid")
                            cellnew.border = brdr
                        else:
                            cellnew.fill = PatternFill(start_color="E5DBBE", end_color="E5DBBE", fill_type="solid")
                            cellnew.border = brdr
        try:
            book.save(cd3file)
            book.close()
        except Exception as e:
            print(str(e))
            print("Exiting!!")
            exit()

    # def backup_file(src_dir, pattern, overwrite):
    def backup_file(src_dir, resource, pattern):
        dest_dir = str(src_dir) + "/backup_" + resource + "/" + datetime.datetime.now().strftime("%d-%m-%H%M%S").replace('/', '-')
        for f in os.listdir(str(src_dir)):
            if f.endswith(pattern):
                print("Backing up existing " + f + " to " + dest_dir)
                if not os.path.exists(dest_dir):
                    # print("\nCreating backup dir " + dest_dir + "\n")
                    os.makedirs(dest_dir)

                src = os.path.join(str(src_dir), f)
                #dest = os.path.join(dest_dir, f)
                # print("backing up ....." + src +"   to  "+dest)
                shutil.move(src, dest_dir)
                """if (overwrite == 'yes'):
                    shutil.move(src, dest_dir)
                elif (overwrite == 'no'):
                    shutil.copyfile(src, dest)
                """


class parseDRGs():
    def __init__(self, filename):
        self.drg_names = {}
        self.drg_rds = {}
        try:
            # Read and search for VCN
            df_drgv2 = pd.read_excel(filename, sheet_name='DRGs', skiprows=1)
        except Exception as e:
            if ("No sheet named" in str(e)):
                print("\nTab - \"DRGs\" is missing in the CD3. Please make sure to use the right CD3 in properties file...Exiting!!")
                exit(1)
            else:
                print("Error occurred while reading the CD3 excel sheet " + str(e))
                exit(1)

        # Drop all empty rows
        df_drgv2 = df_drgv2.dropna(how='all')
        df_drgv2 = df_drgv2.reset_index(drop=True)

        # Create VCN details Dicts and Hub and Spoke VCN Names
        for i in df_drgv2.index:
            region = str(df_drgv2['Region'][i]).strip()
            if region in commonTools.endNames:  # or str(region).lower() == 'nan'):
                break
            drg_name = df_drgv2['DRG Name'][i]
            drg_name = str(drg_name).strip()
            region = region.lower()
            rd_name = str(df_drgv2['Import DRG Route Distribution Name'][i]).strip()

            self.drg_names.setdefault(region, [])

            if (drg_name not in self.drg_names[region]):
                self.drg_names[region].append(drg_name)
                key = drg_name,region
                self.drg_rds.setdefault(key,[])


            if rd_name!='' and rd_name.lower()!='nan':
                self.drg_rds[drg_name,region].append(rd_name)


# NOTE: Does this really need to be a class? its an obsfucated function call.
class parseVCNs():
    def __init__(self, filename):
        self.peering_dict = dict()

        self.vcn_region = {}
        self.vcn_drgs = {}
        self.vcn_drg_names = {}
        self.vcn_compartment = {}
        self.vcn_lpg_names = {}
        self.vcn_lpg_names1 = {}
        self.vcn_lpg_names2 = {}
        self.vcn_lpg_names3 = {}
        self.hub_vcn_names = []
        self.spoke_vcn_names = []
        self.vcn_lpg_rules = {}
        self.vcn_igws = {}
        self.vcn_ngws = {}
        self.vcn_sgws = {}
        self.vcn_hub_spoke_peer_none = {}
        self.vcn_compartment = {}
        self.vcn_names = []
        self.vcn_cidrs = {}
        self.vcns_having_drg = {}

        try:
            # Read and search for VCN
            df_vcn = pd.read_excel(filename, sheet_name='VCNs', skiprows=1)
        except Exception as e:
            if ("No sheet named" in str(e)):
                print("\nTab - \"VCNs\" is missing in the CD3. Please make sure to use the right CD3 in properties file...Exiting!!")
                exit(1)
            else:
                print("Error occurred while reading the CD3 excel sheet " + str(e))
                exit(1)


        # Drop all empty rows
        df_vcn = df_vcn.dropna(how='all')
        df_vcn = df_vcn.reset_index(drop=True)

        # Create VCN details Dicts and Hub and Spoke VCN Names
        for i in df_vcn.index:
            region = str(df_vcn['Region'][i]).strip()
            if region in commonTools.endNames:  # or str(region).lower() == 'nan'):
                break
            vcn_name = df_vcn['VCN Name'][i]
            vcn_name = str(vcn_name).strip()
            region = str(region).strip().lower()

            entry = vcn_name,region
            self.vcn_names.append(entry)

            if str(df_vcn['Hub/Spoke/Peer/None'][i]).strip().split(":")[0].strip().lower() == 'hub':
                self.peering_dict[vcn_name,region] = ""


        for i in df_vcn.index:
            region = str(df_vcn['Region'][i]).strip()
            if (region in commonTools.endNames):  # or str(region).lower()=='nan'):
                break
            vcn_name = df_vcn['VCN Name'][i]
            vcn_name = str(vcn_name).strip()

            region = str(region).strip().lower()
            #self.vcn_region[vcn_name] = region

            self.vcn_lpg_names[vcn_name,region] = str(df_vcn['LPG Required'][i]).strip().split(",")
            self.vcn_lpg_names[vcn_name,region] = [x.strip() for x in self.vcn_lpg_names[vcn_name,region]]

            j = 0
            for lpg in self.vcn_lpg_names[vcn_name,region]:
                if lpg == 'y':
                    self.vcn_lpg_names[vcn_name,region][j] = vcn_name + "_lpg" + str(j)
                    j = j + 1
            self.vcn_lpg_names1[vcn_name,region] = str(df_vcn['LPG Required'][i]).strip().split(",")
            self.vcn_lpg_names1[vcn_name,region] = [x.strip() for x in self.vcn_lpg_names1[vcn_name,region]]

            j = 0
            for lpg in self.vcn_lpg_names1[vcn_name,region]:
                if lpg == 'y':
                    self.vcn_lpg_names1[vcn_name,region][j] = vcn_name + "_lpg" + str(j)
                    j = j + 1

            self.vcn_lpg_names2[vcn_name,region] = str(df_vcn['LPG Required'][i]).strip().split(",")
            self.vcn_lpg_names2[vcn_name,region] = [x.strip() for x in self.vcn_lpg_names2[vcn_name,region]]

            j = 0
            for lpg in self.vcn_lpg_names2[vcn_name,region]:
                if lpg == 'y':
                    self.vcn_lpg_names2[vcn_name,region][j] = vcn_name + "_lpg" + str(j)
                    j = j + 1

            self.vcn_lpg_names3[vcn_name,region] = str(df_vcn['LPG Required'][i]).strip().split(",")
            self.vcn_lpg_names3[vcn_name,region] = [x.strip() for x in self.vcn_lpg_names3[vcn_name,region]]

            j = 0
            for lpg in self.vcn_lpg_names3[vcn_name,region]:
                if lpg == 'y':
                    self.vcn_lpg_names3[vcn_name,region][j] = vcn_name + "_lpg" + str(j)
                    j = j + 1

            self.vcn_drgs[vcn_name,region] = str(df_vcn['DRG Required'][i]).strip()


            if(str(df_vcn['DRG Required'][i]).strip().lower() != 'n'):
                if (str(df_vcn['DRG Required'][i]).strip().lower() == "y"):
                    drg_name = region + "_drg"
                # use name provided in input
                else:
                    drg_name = str(df_vcn['DRG Required'][i]).strip()

                self.vcns_having_drg[vcn_name,region]=drg_name

            self.vcn_cidrs[vcn_name,region]=str(df_vcn['CIDR Blocks'][i]).strip()
            #cidr_blocks = [x.strip() for x in columnvalue.split(',')]

            self.vcn_igws[vcn_name,region] = str(df_vcn['IGW Required'][i]).strip()
            self.vcn_ngws[vcn_name,region] = str(df_vcn['NGW Required'][i]).strip()
            self.vcn_sgws[vcn_name,region] = str(df_vcn['SGW Required'][i]).strip()
            self.vcn_hub_spoke_peer_none[vcn_name,region] = str(df_vcn['Hub/Spoke/Peer/None'][i]).strip().split(":")
            self.vcn_compartment[vcn_name,region] = str(df_vcn['Compartment Name'][i]).strip()
            self.vcn_lpg_rules.setdefault((vcn_name,region), '')

            if (self.vcn_hub_spoke_peer_none[vcn_name,region][0].strip().lower() == 'hub'):
                entry=vcn_name,region
                self.hub_vcn_names.append(entry)

            if (self.vcn_hub_spoke_peer_none[vcn_name,region][0].strip().lower() == 'spoke'):
                hub_name = self.vcn_hub_spoke_peer_none[vcn_name,region][1].strip()
                entry = vcn_name, region
                self.spoke_vcn_names.append(entry)
                try:
                    self.peering_dict[hub_name,region] = self.peering_dict[hub_name,region] + vcn_name + ","
                except KeyError:
                    print("ERROR!!! "+hub_name +" not marked as Hub. Verify hub_spoke_peer_none column again..Exiting!")
                    exit(1)

            if (self.vcn_hub_spoke_peer_none[vcn_name,region][0].strip().lower() == 'peer'):
                self.peering_dict[vcn_name,region] = self.vcn_hub_spoke_peer_none[vcn_name,region][1].strip()


        for k, v in self.peering_dict.items():
            if (v != "" and v[-1] == ','):
                v = v[:-1]
                self.peering_dict[k] = v

@contextmanager
def section(title='', header=False, padding=117):
    separator = '-' if not header else '='
    # Not sure why 117 but thats how it was before.
    print(f'{title:{separator}^{padding}}')
    yield
    if header:
        print(separator * padding)


def exit_menu(msg, exit_code=0):
    print(msg)
    exit(exit_code)


class parseVCNInfo():
    # all_regions = []

    def __init__(self, filename):
        self.onprem_destinations = []
        self.ngw_destinations = []
        self.igw_destinations = []
        try:
            df_info = pd.read_excel(filename, sheet_name='VCN Info', skiprows=1)
        except Exception as e:
            if ("No sheet named" in str(e)):
                print("\nTab - \"VCN Info\" is missing in the CD3. Please make sure to use the right CD3 in properties file...Exiting!!")
                exit(1)
            else:
                print("Error occurred while reading the CD3 excel sheet "+ str(e))
                exit(1)
        # Get Property Values
        values = df_info['Value']

        onprem_destinations = str(values[0]).strip()
        if (onprem_destinations.lower() == 'nan'):
            # print("\ndrg_subnet should not be left empty.. It will create empty route tables")
            self.onprem_destinations.append('')
        else:
            self.onprem_destinations = onprem_destinations.split(",")

        ngw_destinations = str(values[1]).strip()
        if (ngw_destinations.lower() == 'nan'):
            self.ngw_destinations.append('0.0.0.0/0')
        else:
            self.ngw_destinations = ngw_destinations.split(",")

        igw_destinations = str(values[2]).strip()
        if (igw_destinations.lower() == 'nan'):
            self.igw_destinations.append('0.0.0.0/0')
        else:
            self.igw_destinations = igw_destinations.split(",")


class parseSubnets():
    def __init__(self, filename):
        self.vcn_subnet_map = {}
        try:
            # Read and search for VCN
            df_subnet = pd.read_excel(filename, sheet_name='Subnets', skiprows=1)
        except Exception as e:
            if ("No sheet named" in str(e)):
                print("\nTab - \"Subnets\" is missing in the CD3. Please make sure to use the right CD3 in properties file...Exiting!!")
                exit(1)
            else:
                print("Error occurred while reading the CD3 excel sheet " + str(e))
                exit(1)

        # Drop all empty rows
        df_subnet = df_subnet.dropna(how='all')
        df_subnet = df_subnet.reset_index(drop=True)


        for i in df_subnet.index:
            region = str(df_subnet['Region'][i]).strip()
            if (region in commonTools.endNames):  # or str(region).lower()=='nan'):
                break
            key = df_subnet.loc[i,'Region'].strip().lower(),df_subnet.loc[i,'VCN Name'].strip()+"_"+df_subnet.loc[i,'Subnet Name'].strip()
            value = df_subnet.loc[i,'Compartment Name'].strip(), df_subnet.loc[i,'VCN Name'].strip(), df_subnet.loc[i,'Subnet Name'].strip()
            self.vcn_subnet_map[key] =  value

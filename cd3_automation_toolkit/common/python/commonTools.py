
import re
import os
import shutil
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side
import collections
import warnings
warnings.simplefilter("ignore")
from contextlib import contextmanager


@contextmanager
def section(title='', header=False, padding=117):
    separator = '-' if not header else '='
    # Not sure why 117 but thats how it was before.
    print(f'{title:{separator}^{padding}}')
    yield
    if header:
        print(separator * padding)

def exit_menu(msg, exit_code=1):
    print(msg)
    exit(exit_code)

class commonTools():
    endNames = {'<END>', '<end>', '<End>'}

    def __init__(self):
        pass

    # Read rows from CD3
    def data_frame(filename, sheetname):

        # Read the tab from excel, Drop null values, Reset index
        df, col_headers = commonTools.read_cd3(filename, sheetname)
        df = df.dropna(how='all')
        df = df.reset_index(drop=True)

        return df
    def read_cd3(cd3file, sheet_name):
        df = {}
        try:
            df = pd.read_excel(cd3file, sheet_name=sheet_name, skiprows=1, dtype=object)

        except Exception as e:
            if("Events" in str(e) or "Notifications" in str(e)):
                print("\nTabs - \"Events\" or \"Notifications\" is missing in the CD3. Please make sure to use the correct input file for Events and Notifications in properties file...Exiting!!")
                exit(1)
            else:
                print("Error occurred while reading the CD3 excel sheet: "+ str(e))
                exit(1)

        yield df
        try:
            book = load_workbook(cd3file)
            sheet = book[sheet_name]
        except KeyError as e:
            if 'does not exist' in str(e):
                print("\nTab - \""+sheet_name+"\" seems to be missing in the CD3. Please make sure to use the right CD3 in properties file.....Exiting!!")
                exit(1)
        except Exception as e:
            print(str(e))
            print("Exiting!!")
            exit(1)

        values_for_column = collections.OrderedDict()
        # values_for_column={}
        for j in range(0, sheet.max_column):
            col_name = sheet.cell(row=2, column=j + 1).value
            if (type(col_name) == str):
                values_for_column[col_name] = []
        yield values_for_column

    #Write exported  rows to cd3
    def write_to_cd3(values_for_column, cd3file, sheet_name,append=False):
        try:
            book = load_workbook(cd3file)
            sheet = book[sheet_name]


        except Exception as e:
            print(str(e))
            print("Exiting!!")
            exit(1)
        if (sheet_name == "VCN Info"):
            onprem_destinations = ""
            ngw_destinations = ""
            igw_destinations = ""
            for destination in values_for_column["onprem_destinations"]:
                onprem_destinations=destination+","+onprem_destinations
            for destination in values_for_column["ngw_destinations"]:
                ngw_destinations = destination + "," + ngw_destinations
            for destination in values_for_column["igw_destinations"]:
                igw_destinations = destination + "," + igw_destinations

            if (onprem_destinations != "" and onprem_destinations[-1] == ','):
                onprem_destinations = onprem_destinations[:-1]
            if (ngw_destinations != "" and ngw_destinations[-1] == ','):
                ngw_destinations = ngw_destinations[:-1]
            if (igw_destinations != "" and igw_destinations[-1] == ','):
                igw_destinations = igw_destinations[:-1]

            sheet.cell(3,2).value = onprem_destinations
            sheet.cell(4,2).value = ngw_destinations
            sheet.cell(5,2).value = igw_destinations
            try:
                book.save(cd3file)
                book.close()
            except Exception as e:
                print(str(e))
                print("Exiting!!")
                exit(1)
            return


        #rows_len=len(rows)
        rows_len = len(values_for_column["Region"])
        sheet_max_rows = sheet.max_row
        #If no rows exported from OCI, remove the sample data as well
        if(rows_len == 0) :
            if not append:
                print("0 rows exported; Nothing to write to CD3 excel; Tab "+sheet_name +" will be empty in CD3 excel!!")
                for i in range(0, sheet.max_row):
                    for j in range(0, sheet.max_column):
                        sheet.cell(row=i + 3, column=j + 1).value = ""
                try:
                    book.save(cd3file)
                    book.close()
                except Exception as e:
                    print(str(e))
                    print("Exiting!!")
                    exit(1)
            return

        if append:
            for x in range(1, sheet_max_rows):
                if sheet['A'][x].value == None:
                    last_line = x
                    break
            #rows_len +=last_line
            large = rows_len
            start = last_line+1


        else:
            start = 3
            if (rows_len > sheet_max_rows):
                large = rows_len
            else:
                large = sheet_max_rows

        df, values_for_column_sheet = commonTools.read_cd3(cd3file, sheet_name)

        #Put Data
        j=0
        for i in range(0,large):

            for col_name in values_for_column.keys():

                #Check if column name to be populated in present in the sheet.
                if col_name not in values_for_column_sheet:
                    continue
                # Data
                if(i>=rows_len):
                    sheet.cell(row=i+start, column=j+1).value = ""
                else:
                    sheet.cell(row=i+start, column=j+1).value = values_for_column[col_name][i]
                sheet.cell(row=i+start, column=j+1).alignment = Alignment(wrap_text=True)
                j=j+1
            j=0


        brdr = Border(left=Side(style='thin'),
                      right=Side(style='thin'),
                      top=Side(style='thin'),
                      bottom=Side(style='thin'),
                      )

        for row in sheet.iter_rows(min_row=3):
            for cell in row:
                cell.border = brdr

        # Add color for exported sec rules and route rules
        if (sheet_name == "RouteRulesinOCI" or sheet_name == "SecRulesinOCI" or sheet_name == "DRGRouteRulesinOCI"):
            names = []
            # Add color coding to exported rules
            for row in sheet.iter_rows(min_row=3):
                c = 0
                region = ""
                name = ""
                for cell in row:
                    c = c + 1
                    if (c == 1):
                        region = cell.value
                        continue
                    elif (c == 4):
                        name = cell.value
                        break

                vcn_name = region + "_" + name
                if (vcn_name not in names):
                    names.append(vcn_name)
                    for cellnew in row:
                        if (len(names) % 2 == 0):
                            cellnew.fill = PatternFill(start_color="94AFAF", end_color="94AFAF", fill_type="solid")
                            cellnew.border = brdr
                        else:
                            cellnew.fill = PatternFill(start_color="E5DBBE", end_color="E5DBBE", fill_type="solid")
                            cellnew.border = brdr
                else:
                    for cellnew in row:
                        if (len(names) % 2 == 0):
                            cellnew.fill = PatternFill(start_color="94AFAF", end_color="94AFAF", fill_type="solid")
                            cellnew.border = brdr
                        else:
                            cellnew.fill = PatternFill(start_color="E5DBBE", end_color="E5DBBE", fill_type="solid")
                            cellnew.border = brdr
        try:
            book.save(cd3file)
            book.close()
        except Exception as e:
            print(str(e))
            print("Exiting!!")
            exit(1)

            # Check value exported
            # If None - replace with ""
            # If list, convert to comma sepearted string
    def check_exported_value(value):
            if value == None:
                value = ""
            if ("list" in str(type(value))):
                str1 = ""
                if (value.__len__() == 0):
                    value = ""
                for v in value:
                    str1 = v + "," + str1
                if (str1 != "" and str1[-1] == ','):
                    value = str1[:-1]

            return value
        # Check TF variable Name
    def check_tf_variable(var_name):
            tfname = re.compile('[^a-zA-Z0-9_-]')
            tfnamestart = re.compile('[A-Za-z]')

            var_name = tfname.sub("-", var_name)
            x = tfnamestart.match(var_name)
            # variable name doesnot start with letter; append with c
            if (x == None):
                var_name = "c" + var_name
            return var_name

    # Process ColumnValues
    def check_columnvalue(columnvalue):

        if str(columnvalue).lower() == 'true' or str(columnvalue).lower() == 'false':
            columnvalue = str(columnvalue).lower()

        if (columnvalue.lower() == 'nan'):
            columnvalue = ""

        # replace \ with \\
        if("\\" in columnvalue):
            columnvalue = columnvalue.replace("\\", "\\\\")

        # replace " with \"
        if("\"" in columnvalue):
            columnvalue=columnvalue.replace("\"","\\\"")

        return columnvalue

    # Process column values with ::
    def check_multivalues_columnvalue(columnvalue, columnname, tempdict):
        columnvalue = str(columnvalue).strip()
        columnname = commonTools.check_column_headers(columnname)
        if "::" in columnvalue:
            if ".Flex" in columnvalue or ".Micro" in columnvalue:
                columnname = commonTools.check_column_headers(columnname)
                multivalues = columnvalue.split("::")
                multivalues = [str(part).strip() for part in multivalues if part]
                tempdict = {columnname: multivalues}
            elif columnname != 'Compartment Name' and "ipv6" not in columnname.lower():
                columnname = commonTools.check_column_headers(columnname)
                multivalues = columnvalue.split("::")
                multivalues = [str(part).strip() for part in multivalues ]#if part]
                tempdict = {columnname: multivalues}
        return tempdict

    # Check CD3 Column headers
    def check_column_headers(var_name):
        # replace special characters and spaces with '_' and convert to lowercase
        # replaces multiple occurrence of '_' to just 1
        var_name = var_name.strip()
        var_name = re.sub('[@!#$%^&*<>?/}{~: \n()|-]', '_', var_name).lower()
        var_name = re.sub('_+', '_', var_name).lower()
        return var_name

    def backup_file(src_dir, resource, pattern):
        dest_dir = str(src_dir) + "/backup_" + resource + "/" + datetime.datetime.now().strftime("%d-%m-%H%M%S").replace('/', '-')
        for f in os.listdir(str(src_dir)):
            if f.endswith(pattern):
                print("Backing up existing " + f + " to " + dest_dir)
                if not os.path.exists(dest_dir):
                    # print("\nCreating backup dir " + dest_dir + "\n")
                    os.makedirs(dest_dir)

                src = os.path.join(str(src_dir), f)
                #dest = os.path.join(dest_dir, f)
                # print("backing up ....." + src +"   to  "+dest)
                shutil.move(src, dest_dir)
                """if (overwrite == 'yes'):
                    shutil.move(src, dest_dir)
                elif (overwrite == 'no'):
                    shutil.copyfile(src, dest)
                """

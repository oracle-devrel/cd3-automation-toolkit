#!/bin/python


import argparse
import sys
import oci
from oci.core.virtual_network_client import VirtualNetworkClient
from oci.identity import IdentityClient
import pandas as pd
from openpyxl import load_workbook

def convertNullToNothing(input):
    EMPTY_STRING = ""
    if input is None:
        return EMPTY_STRING
    else:
        return str(input)

compartment_ids={}
def get_network_compartment_id(config):#, compartment_name):
    identity = IdentityClient(config)
    comp_list = oci.pagination.list_call_get_all_results(identity.list_compartments,config["tenancy"],compartment_id_in_subtree=True)
    compartment_list = comp_list.data

    #if(compartment_name=='root'):
    for compartment in compartment_list:
        if(compartment.lifecycle_state == 'ACTIVE'):
            compartment_ids[compartment.name]=compartment.id
    return compartment_ids

    '''else:
        for compartment in compartment_list:
            if compartment.name == compartment_name:
                compartment_ids[compartment.name]=compartment.id
                return compartment_ids
'''
def get_vcn_id(config,compartment_id,vname):
    vcncient = VirtualNetworkClient(config)
    vcnlist = vcncient.list_vcns(compartment_id=compartment_id,lifecycle_state="AVAILABLE")
    vcn_name = vname.lower()
    for v in vcnlist.data:
        name = v.display_name
        if name.lower() == vcn_name:
            return v.id

def get_vcns(config,compartment_id):
    vcncient = VirtualNetworkClient(config)
    vcnlist = vcncient.list_vcns(compartment_id=compartment_id,lifecycle_state="AVAILABLE")
    return vcnlist


def print_secrules(seclists,region,vcn_name,comp_name):
    #print ("SubnetName, RuleType, Protocol, isStateless, Source, SPortMin, SPortMax, Destination, DPortMin, DPortMax, ICMPType, ICMPCode")
    #oname.write("SubnetName, RuleType, Protocol, isStateless, Source, SPortMin, SPortMax, Destination, DPortMin, DPortMax, ICMPType, ICMPCode\n")

    global i
    global df
    for seclist in seclists.data:
        isec_rules = seclist.ingress_security_rules
        esec_rules = seclist.egress_security_rules
        display_name = seclist.display_name

        # display name contaoin AD1, AD2 or AD3 and CIDR
        if ('-ad1-10.' in display_name or '-ad2-10.' in display_name or '-ad3-10.' in display_name or '-ad1-172.' in display_name
                or '-ad2-172.' in display_name or '-ad3-172.' in display_name or '-ad1-192.' in display_name or '-ad2-192.' in display_name
                or '-ad3-192.' in display_name):
            dn = display_name.rsplit("-", 2)[0]

        # display name contains CIDR
        elif ('-10.' in display_name or '-172.' in display_name or '192.' in display_name):
            dn = display_name.rsplit("-", 1)[0]
        else:
            dn = display_name#.rsplit("-", 1)[0]
        if(len(isec_rules)==0 and len(esec_rules)==0):
            new_row = pd.DataFrame(
                {'SubnetName': dn, 'RuleType': '', 'Protocol': '', 'isStateless':'',
                 'Source': '', 'SPortMin': '', 'SPortMax': '', 'Destination': '', 'DPortMin': '',
                 'DPortMax': '',
                 'ICMPType': '', 'ICMPCode': '', 'VCN Name': vcn_name, 'Compartment Name': comp_name,'Region':region}, index=[i])
            df = df.append(new_row, ignore_index=True)
        for rule in esec_rules:
            if rule.protocol == "all":
                print(dn + ",egress,all," + str(rule.is_stateless) + "," + rule.destination + ",,,,,,,")
                new_row = pd.DataFrame(
                    {'SubnetName': dn, 'RuleType': 'egress', 'Protocol': 'all', 'isStateless': str(rule.is_stateless),
                     'Source': '', 'SPortMin': '', 'SPortMax': '', 'Destination': rule.destination, 'DPortMin': '',
                     'DPortMax': '','ICMPType': '', 'ICMPCode': '', 'VCN Name': vcn_name, 'Compartment Name': comp_name,'Region':region}, index=[i])
                df = df.append(new_row, ignore_index=True)

        for rule in isec_rules:
          #  print (rule)
            i+=1
            print(i)
            if rule.protocol == "6":
                if rule.tcp_options is None:
                    print (dn + ",ingress,tcp," + str(rule.is_stateless) + "," + rule.source + ",,,,,,,")
                    #oname.write(dn + ",ingress,tcp," + str(rule.is_stateless) + "," + rule.source + ",,,,,,,\n")

                    new_row = pd.DataFrame({'SubnetName':dn,'RuleType':'ingress','Protocol':'tcp','isStateless':str(rule.is_stateless),
                                            'Source':rule.source,'SPortMin':'','SPortMax':'','Destination':'','DPortMin':'','DPortMax':'',
                                            'ICMPType':'','ICMPCode':'','VCN Name':vcn_name,'Compartment Name':comp_name,'Region':region},index =[i])
                #else:
                #    print(rule.tcp_options.destination_port_range)
                elif rule.tcp_options.destination_port_range is not None:
                    min = convertNullToNothing(rule.tcp_options.destination_port_range.min)
                    max = convertNullToNothing(rule.tcp_options.destination_port_range.max)
                    print (dn + ",ingress,tcp," + str(rule.is_stateless) + "," + rule.source + ",,,," + min + "," + max+",,")
                    #oname.write(dn + ",ingress,tcp," + str(rule.is_stateless) + "," + rule.source + ",,,," + min + "," + max+",,\n")
                    new_row = pd.DataFrame({'SubnetName': dn, 'RuleType': 'ingress', 'Protocol': 'tcp',
                                            'isStateless': str(rule.is_stateless),
                                            'Source': rule.source, 'SPortMin': '', 'SPortMax': '', 'Destination': '',
                                            'DPortMin': min, 'DPortMax': max,
                                            'ICMPType': '', 'ICMPCode': '','VCN Name':vcn_name,'Compartment Name':comp_name,'Region':region},index =[i])
                else:
                    new_row = pd.DataFrame({'SubnetName': dn, 'RuleType': 'ingress', 'Protocol': 'tcp',
                                            'isStateless': str(rule.is_stateless),
                                            'Source': rule.source, 'SPortMin': '', 'SPortMax': '', 'Destination': '',
                                            'DPortMin': '', 'DPortMax': '',
                                            'ICMPType': '', 'ICMPCode': '', 'VCN Name': vcn_name,
                                            'Compartment Name': comp_name, 'Region': region}, index=[i])


            if rule.protocol == "1":
                if rule.icmp_options is None:
                    print (dn + ",ingress,icmp," + str(rule.is_stateless) + "," + rule.source + ",,,,,,,")
                    #oname.write(dn + ",ingress,icmp," + str(rule.is_stateless) + "," + rule.source + ",,,,,,,\n")
                    new_row = pd.DataFrame({'SubnetName': dn, 'RuleType': 'ingress', 'Protocol': 'icmp',
                                            'isStateless': str(rule.is_stateless),
                                            'Source': rule.source, 'SPortMin': '', 'SPortMax': '', 'Destination': '',
                                            'DPortMin': '', 'DPortMax': '',
                                            'ICMPType': '', 'ICMPCode': '','VCN Name':vcn_name,'Compartment Name':comp_name,'Region':region},index =[i])
                else:
                    code = convertNullToNothing(rule.icmp_options.code)
                    type = convertNullToNothing(rule.icmp_options.type)
                    print (dn + ",ingress,icmp," + str(rule.is_stateless) + "," + rule.source + ",,,,,," + type + "," + code)
                    #oname.write(dn + ",ingress,icmp," + str(rule.is_stateless) + "," + rule.source + ",,,,,," + type + "," + code+"\n")
                    new_row = pd.DataFrame(
                    {'SubnetName': dn, 'RuleType': 'ingress', 'Protocol': 'icmp', 'isStateless': str(rule.is_stateless),
                     'Source': rule.source, 'SPortMin': '', 'SPortMax': '', 'Destination': '', 'DPortMin': '',
                     'DPortMax': '',
                     'ICMPType': type, 'ICMPCode': code,'VCN Name':vcn_name,'Compartment Name':comp_name,'Region':region},index =[i])

            if rule.protocol == "17":
                if rule.udp_options is None:
                    print (dn + ",ingress,udp," + str(rule.is_stateless) + "," + rule.source + ",,,,,,,")
                    #oname.write(dn + ",ingress,udp," + str(rule.is_stateless) + "," + rule.source + ",,,,,,,\n")
                    new_row = pd.DataFrame({'SubnetName': dn, 'RuleType': 'ingress', 'Protocol': 'udp',
                                            'isStateless': str(rule.is_stateless),
                                            'Source': rule.source, 'SPortMin': '', 'SPortMax': '', 'Destination': '',
                                            'DPortMin': '', 'DPortMax': '',
                                            'ICMPType': '', 'ICMPCode': '','VCN Name':vcn_name,'Compartment Name':comp_name,'Region':region},index =[i])
                elif rule.udp_options.destination_port_range is not None:
                    min = convertNullToNothing(rule.udp_options.destination_port_range.min)
                    max = convertNullToNothing(rule.udp_options.destination_port_range.max)
                    print (dn + ",ingress,udp," + str(rule.is_stateless) + "," + rule.source + ",,,," + min + "," + max+",,")
                    #oname.write(dn + ",ingress,udp," + str(rule.is_stateless) + "," + rule.source + ",,,," + min + "," + max+",,\n")
                    new_row = pd.DataFrame({'SubnetName': dn, 'RuleType': 'ingress', 'Protocol': 'udp',
                                            'isStateless': str(rule.is_stateless),
                                            'Source': rule.source, 'SPortMin': '', 'SPortMax': '', 'Destination': '',
                                            'DPortMin': min, 'DPortMax': max,
                                            'ICMPType': '', 'ICMPCode': '','VCN Name':vcn_name,'Compartment Name':comp_name,'Region':region},index =[i])
                else:
                    new_row = pd.DataFrame({'SubnetName': dn, 'RuleType': 'ingress', 'Protocol': 'udp',
                                            'isStateless': str(rule.is_stateless),
                                            'Source': rule.source, 'SPortMin': '', 'SPortMax': '', 'Destination': '',
                                            'DPortMin': '', 'DPortMax': '',
                                            'ICMPType': '', 'ICMPCode': '', 'VCN Name': vcn_name,
                                            'Compartment Name': comp_name, 'Region': region}, index=[i])
            if rule.protocol == "all":
                print (dn + ",ingress,all," + str(rule.is_stateless) + "," + rule.source + ",,,,,,,")
                #oname.write(dn + ",ingress,all," + str(rule.is_stateless) + "," + rule.source + ",,,,,,,\n")
                new_row = pd.DataFrame(
                    {'SubnetName': dn, 'RuleType': 'ingress', 'Protocol': 'all', 'isStateless': str(rule.is_stateless),
                     'Source': rule.source, 'SPortMin': '', 'SPortMax': '', 'Destination': '', 'DPortMin': '',
                     'DPortMax': '',
                     'ICMPType': '', 'ICMPCode': '','VCN Name':vcn_name,'Compartment Name':comp_name,'Region':region},index =[i])
            #df = pd.concat([new_row, df],ignore_index =True)

            df=df.append(new_row,ignore_index =True)
        print("----------------------------------------------")


parser = argparse.ArgumentParser(description="Export Security list on OCI to CD3")
parser.add_argument("cd3file", help="path of CD3 excel file to export rules to")
parser.add_argument("--vcnName", help="VCN from which to export the sec list", required=False)
parser.add_argument("--networkCompartment", help="Compartment where VCN resides", required=False)
parser.add_argument("--configFileName", help="Config file name" , required=False)



if len(sys.argv) < 2:
    parser.print_help()
    sys.exit(1)

args = parser.parse_args()
cd3file=args.cd3file

if('.xls' not in cd3file):
    print("\nAcceptable cd3 format: .xlsx")
    exit()

vcn_name = args.vcnName

if args.configFileName is not None:
    configFileName = args.configFileName
    config = oci.config.from_file(file_location=configFileName)
else:
    config = oci.config.from_file()

ntk_compartment_ids = get_network_compartment_id(config)#, ntk_comp_name)

i=0
df=pd.DataFrame()

if vcn_name is not None:
    ntk_comp_name = args.networkCompartment
    if(ntk_comp_name=='' or ntk_comp_name is None):
        print("\nEnter a valid name for the compartment weher VCN resides...")
        exit(1)

    found = ''
    print("\nSearching for VCN in ASH region...")
    config.__setitem__("region", "us-ashburn-1")
    vcn_ocid = get_vcn_id(config, ntk_compartment_ids[ntk_comp_name], vcn_name)

    if (vcn_ocid == '' or vcn_ocid is None):
        print('\nCould not find vcn in this compartment in ASH region...Searching in PHX region...')
        config.__setitem__("region", "us-phoenix-1")
        vcn_ocid = get_vcn_id(config, ntk_compartment_ids[ntk_comp_name], vcn_name)
        if (vcn_ocid == '' or vcn_ocid is None):
            print('\nCould not find vcn in this compartment in PHX region also..verify the vcn name and compartment name where vcn resides and try again...')
            exit(1)
        else:
            print("\nFound in Phoenix...")
            found = 'phx'
    else:
        print("\nFound in Ashburn...")
        found = 'ash'

    if (found == 'ash'):
        config.__setitem__("region", "us-ashburn-1")
        vcn = VirtualNetworkClient(config)
        region = 'Ashburn'
    if (found == 'phx'):
        config.__setitem__("region", "us-phoenix-1")
        vcn = VirtualNetworkClient(config)
        region = 'Phoenix'

    seclists = vcn.list_security_lists(compartment_id=ntk_compartment_ids[ntk_comp_name], vcn_id=vcn_ocid, lifecycle_state='AVAILABLE',sort_by='DISPLAYNAME')
    print_secrules(seclists,region,vcn_name,ntk_comp_name)
else:
    print("\nFetching Security Rules for all VCNs in tenancy...")
    for ntk_compartment_name in ntk_compartment_ids:
        config.__setitem__("region", "us-ashburn-1")
        vcn = VirtualNetworkClient(config)
        vcns_ash = get_vcns(config, ntk_compartment_ids[ntk_compartment_name])
        for v in vcns_ash.data:
            vcn_id = v.id
            vcn_name=v.display_name
            seclists = vcn.list_security_lists(compartment_id=ntk_compartment_ids[ntk_compartment_name], vcn_id=vcn_id, lifecycle_state='AVAILABLE',sort_by='DISPLAYNAME')
            print_secrules(seclists,'Ashburn',vcn_name,ntk_compartment_name)

        config.__setitem__("region", "us-phoenix-1")
        vcn = VirtualNetworkClient(config)
        vcns_phx = get_vcns(config, ntk_compartment_ids[ntk_compartment_name])
        for v in vcns_phx.data:
            vcn_id = v.id
            vcn_name = v.display_name
            seclists = vcn.list_security_lists(compartment_id=ntk_compartment_ids[ntk_compartment_name], vcn_id=vcn_id,lifecycle_state='AVAILABLE',sort_by='DISPLAYNAME')
            print_secrules(seclists, 'Phoenix', vcn_name, ntk_compartment_name)

#oname.close()

book = load_workbook(cd3file)
if('SecRulesinOCI' in book.sheetnames):
    book.remove(book['SecRulesinOCI'])

writer = pd.ExcelWriter(cd3file, engine='openpyxl')
writer.book = book
writer.save()

book = load_workbook(cd3file)
writer = pd.ExcelWriter(cd3file, engine='openpyxl')
writer.book = book
df.to_excel(writer, sheet_name='SecRulesinOCI', index=False)
writer.save()

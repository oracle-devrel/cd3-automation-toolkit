#!/bin/python


import argparse
import sys
import oci
from oci.core.virtual_network_client import VirtualNetworkClient
from oci.identity import IdentityClient
import pandas as pd
from openpyxl import load_workbook

def convertNullToNothing(input):
    EMPTY_STRING = ""
    if input is None:
        return EMPTY_STRING
    else:
        return str(input)

compartment_ids={}
def get_network_compartment_id(config):#, compartment_name):
    identity = IdentityClient(config)
    comp_list = oci.pagination.list_call_get_all_results(identity.list_compartments,config["tenancy"],compartment_id_in_subtree=True)
    compartment_list = comp_list.data

    #if(compartment_name=='root'):
    for compartment in compartment_list:
        if(compartment.lifecycle_state == 'ACTIVE'):
            compartment_ids[compartment.name]=compartment.id
    return compartment_ids

    '''else:
        for compartment in compartment_list:
            if compartment.name == compartment_name:
                compartment_ids[compartment.name]=compartment.id
                return compartment_ids

'''
def get_vcn_id(config,compartment_id,vname):
    vcncient = VirtualNetworkClient(config)
    vcnlist = vcncient.list_vcns(compartment_id=compartment_id,lifecycle_state="AVAILABLE")
    vcn_name = vname.lower()
    for v in vcnlist.data:
        name = v.display_name
        if name.lower() == vcn_name:
            return v.id

def get_vcns(config,compartment_id):
    vcncient = VirtualNetworkClient(config)
    vcnlist = vcncient.list_vcns(compartment_id=compartment_id,lifecycle_state="AVAILABLE")
    return vcnlist

def print_routetables(routetables,region,vcn_name,comp_name):

    #oname.write("SubnetName, Destination CIDR, Route Destination Object, Destination Type\n")
    global i
    global df
    for routetable in routetables.data:
        print(routetable.display_name)
        rules = routetable.route_rules
        display_name = routetable.display_name

        # display name contains AD1, AD2 or AD3 and CIDR
        if ('-ad1-10.' in display_name or '-ad2-10.' in display_name or '-ad3-10.' in display_name or '-ad1-172.' in display_name
                or '-ad2-172.' in display_name or '-ad3-172.' in display_name or '-ad1-192.' in display_name or '-ad2-192.' in display_name
                or '-ad3-192.' in display_name):
            dn = display_name.rsplit("-", 2)[0]

        # display name contains CIDR
        elif ('-10.' in display_name or '-172.' in display_name or '192.' in display_name):
            dn = display_name.rsplit("-", 1)[0]
        else:
            dn=display_name

        #dn = routetable.display_name
        if(not rules):
            i=i+1
            print(i)
            new_row = pd.DataFrame({'SubnetName': dn, 'Destination CIDR': '',
                                    'Route Destination Object': '',
                                    'Destination Type': '', 'VCN Name':vcn_name,'Compartment Name':comp_name,'Region':region}, index=[i])
            df = df.append(new_row, ignore_index=True)
        for rule in rules:
            i=i+1
            print(i)
            print(dn + "," + str(rule.destination) + "," + str(rule.network_entity_id)+","+ str(rule.destination_type))
            #oname.write(dn + "," + str(rule.destination) + "," +str(rule.network_entity_id)+","+ str(rule.destination_type)+"\n")
            new_row = pd.DataFrame({'SubnetName': dn, 'Destination CIDR': str(rule.destination), 'Route Destination Object': str(rule.network_entity_id), 'Destination Type': str(rule.destination_type),'VCN Name':vcn_name,'Compartment Name':comp_name,'Region':region}, index=[i])
            df = df.append(new_row, ignore_index=True)

parser = argparse.ArgumentParser(description="Export Route Table on OCI to CD3")
parser.add_argument("cd3file", help="path of CD3 excel file to export rules to")
parser.add_argument("--vcnName", help="VCN from which to export the sec list; Leave blank if need to export from all VCNs", required=False)
parser.add_argument("--networkCompartment", help="Compartment where VCN resides", required=False)
parser.add_argument("--configFileName", help="Config file name" , required=False)



if len(sys.argv) < 2:
    parser.print_help()
    sys.exit(1)

args = parser.parse_args()
cd3file=args.cd3file

if('.xls' not in cd3file):
    print("\nAcceptable cd3 format: .xlsx")
    exit()

vcn_name = args.vcnName

if args.configFileName is not None:
    configFileName = args.configFileName
    config = oci.config.from_file(file_location=configFileName)
else:
    config = oci.config.from_file()


ntk_compartment_ids = get_network_compartment_id(config)#, ntk_comp_name)

i=0
df=pd.DataFrame()

if vcn_name is not None:
    ntk_comp_name = args.networkCompartment
    if(ntk_comp_name=='' or ntk_comp_name is None):
        print("\nEnter a valid name for the compartment where VCN resides...")
        exit(1)

    found = ''
    print("\nSearching for VCN in ASH region...")
    config.__setitem__("region", "us-ashburn-1")
    vcn_ocid = get_vcn_id(config,ntk_compartment_ids[ntk_comp_name],vcn_name)

    if(vcn_ocid=='' or vcn_ocid is None):
        print('\nCould not find vcn in this compartment in ASH region...Searching in PHX region...')
        config.__setitem__("region", "us-phoenix-1")
        vcn_ocid = get_vcn_id(config, ntk_compartment_ids[ntk_comp_name], vcn_name)
        if (vcn_ocid == '' or vcn_ocid is None):
            print('\nCould not find vcn in this compartment in PHX region also..verify the vcn name and compartment name where vcn resides and try again...')
            exit(1)
        else:
            print("\nFound in Phoenix...")
            found='phx'
    else:
        print("\nFound in Ashburn...")
        found='ash'

    if(found=='ash'):
        config.__setitem__("region", "us-ashburn-1")
        vcn = VirtualNetworkClient(config)
        region='Ashburn'
    if(found=='phx'):
        config.__setitem__("region", "us-phoenix-1")
        vcn = VirtualNetworkClient(config)
        region='Phoenix'

    routetables = vcn.list_route_tables(compartment_id=ntk_compartment_ids[ntk_comp_name], vcn_id=vcn_ocid, lifecycle_state='AVAILABLE')
    print_routetables(routetables,region, vcn_name,ntk_comp_name)
else:
    print("\nFetching for all VCNs in tenancy...")
    for ntk_compartment_name in ntk_compartment_ids:
        config.__setitem__("region", "us-ashburn-1")
        vcn = VirtualNetworkClient(config)
        vcns_ash = get_vcns(config,ntk_compartment_ids[ntk_compartment_name])
        for v in vcns_ash.data:
            vcn_id = v.id
            vcn_name=v.display_name
            routetables = vcn.list_route_tables(compartment_id=ntk_compartment_ids[ntk_compartment_name], vcn_id=vcn_id, lifecycle_state='AVAILABLE')
            print_routetables(routetables,'Ashburn',vcn_name,ntk_compartment_name)

        config.__setitem__("region", "us-phoenix-1")
        vcn = VirtualNetworkClient(config)
        vcns_phx = get_vcns(config, ntk_compartment_ids[ntk_compartment_name])
        for v in vcns_phx.data:
            vcn_id = v.id
            vcn_name = v.display_name
            routetables = vcn.list_route_tables(compartment_id=ntk_compartment_ids[ntk_compartment_name], vcn_id=vcn_id,
                                                lifecycle_state='AVAILABLE')
            print_routetables(routetables,'Phoenix',vcn_name, ntk_compartment_name)

#oname.close()

book = load_workbook(cd3file)
if('RouteRulesinOCI' in book.sheetnames):
    book.remove(book['RouteRulesinOCI'])
writer = pd.ExcelWriter(cd3file, engine='openpyxl')
writer.book = book
writer.save()

book = load_workbook(cd3file)
writer = pd.ExcelWriter(cd3file, engine='openpyxl')
writer.book = book
df.to_excel(writer, sheet_name='RouteRulesinOCI', index=False)
writer.save()

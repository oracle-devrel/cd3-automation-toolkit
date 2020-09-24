import pandas as pd
import os
import shutil
import datetime
import configparser
import oci
from oci.identity import IdentityClient
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side
import collections
import re
import json as simplejson

class commonTools():
    all_regions=[]
    home_region=""
    ntk_compartment_ids = {}
    region_dict={}
    protocol_dict={}
    sheet_dict={}
    endNames = {'<END>', '<end>', '<End>'}
    tagColumns = {'freeform tags', 'freeform_tags', 'defined_tags', 'defined tags'}

    #Read Regions and Protocols Files and Excel_Columns and create dicts
    def __init__(self):
        #When called from wthin BaseNetwork
        dir=os.getcwd()
        if("ExportFromOCI" in dir):
            os.chdir("../")
        elif ("Solutions" in dir):
            os.chdir("../../")
        if("Networking" in dir):
            os.chdir("../../../")
        elif ("CoreInfra" in dir):
            os.chdir("../../")
        elif ("Identity" in dir):
            os.chdir("../../")
        elif ("Database" in dir):
            os.chdir("../")
        elif ("Governance" in dir):
            os.chdir("../../")
        elif ("OCSWorkVM" in dir):
            os.chdir("../")
        regionFileName="OCI_Regions"
        protocolFileName="OCI_Protocols"
        excelColumnName="Excel_Columns"
        with open (regionFileName) as f:
            for line in f:
                key = line.split(":")[0].strip()
                val = line.split(":")[1].strip()
                self.region_dict[key] = val

        with open(protocolFileName) as f:
            for line in f:
                key=line.split(":")[0].strip()
                val=line.split(":")[1].strip()
                self.protocol_dict[key]=val

        #Get Dict for column names
        with open(excelColumnName) as f:
            s = f.read()
        while True:
            decoder = simplejson.JSONDecoder()
            obj, end = decoder.raw_decode(s)
            self.sheet_dict = dict(obj)
            break

        #Change back to Initial
        os.chdir(dir)

    #Get Tenancy Regions
    def get_subscribedregions(self,configFileName):
        #Get config client
        if configFileName=="":
            config = oci.config.from_file()
        else:
            config = oci.config.from_file(file_location=configFileName)
        idc = IdentityClient(config,retry_strategy=oci.retry.DEFAULT_RETRY_STRATEGY)
        regionsubscriptions = idc.list_region_subscriptions(tenancy_id=config['tenancy'])
        homeregion=""
        for rs in regionsubscriptions.data:
            if (rs.is_home_region == True):
                homeregion = rs.region_name
            for k, v in self.region_dict.items():
                if (homeregion!="" and v == homeregion):
                    self.home_region = k
                if (rs.region_name == v):
                    self.all_regions.append(k)

        del config
        del idc

    #Get Compartment OCIDs
    def get_network_compartment_ids(self,c_id, c_name,configFileName):
        # Get config client
        if configFileName == "":
            config = oci.config.from_file()
        else:
            config = oci.config.from_file(file_location=configFileName)

        tenancy_id=config['tenancy']
        idc = IdentityClient(config,retry_strategy=oci.retry.DEFAULT_RETRY_STRATEGY)
        compartments = oci.pagination.list_call_get_all_results(idc.list_compartments,compartment_id=c_id, compartment_id_in_subtree=False)

        for c in compartments.data:
            if c.lifecycle_state == "ACTIVE":
                if (c_name != "root"):
                    name = c_name + "::" + c.name
                else:
                    name = c.name
                self.ntk_compartment_ids[name] = c.id
                # Put individual compartment names also in the dictionary
                if (name != c.name):
                    if c.name not in self.ntk_compartment_ids.keys():
                        self.ntk_compartment_ids[c.name] = c.id
                    else:
                        # Remove the individual name added to dict as it is duplicate only if its parent comp is not root
                        c_details = idc.get_compartment(self.ntk_compartment_ids[c.name]).data
                        if (c_details.compartment_id != tenancy_id):
                            self.ntk_compartment_ids.pop(c.name)

                self.get_network_compartment_ids(c.id, name,configFileName)

        self.ntk_compartment_ids["root"]=tenancy_id
        del tenancy_id
        del idc
        del config

    #Check value exported
    #If None - replace with ""
    #If list, convert to comma sepearted string
    def check_exported_value(value):
        if value == None:
            value = ""
        if ("list" in str(type(value))):
            str1 = ""
            if(value.__len__()==0):
                value=""
            for v in value:
                str1 = v + "," + str1
            if (str1 != "" and str1[-1] == ','):
                value = str1[:-1]

        return value

    # Export Tag fields - common code - Defined and Freeform Tags
    # header - individual headers/column name
    # values_for_column - list of columns from read_cd3 function
    def export_tags(oci_obj, header, values_for_column):
        defined_tags=""
        freeform_tags=""
        if 'defined' in str(header).lower():
            if (oci_obj.__getattribute__('defined_tags')):
                for namespace, tags in oci_obj.__getattribute__('defined_tags').items():
                    for key, value in tags.items():
                        # Each Namespace/TagKey - Value pair ends with a ;
                        value = str(namespace + "." + key + "=" + value)
                        defined_tags=value+","+defined_tags

            if (defined_tags != "" and defined_tags[-1] == ','):
                defined_tags = defined_tags[:-1]
            values_for_column[header].append(defined_tags)

        elif 'free' in str(header).lower():
            if (oci_obj.__getattribute__('freeform_tags')):
                for keys, values in oci_obj.__getattribute__('freeform_tags').items():
                    value = str(keys + "=" + values)
                    freeform_tags= value + ',' +freeform_tags

            if (freeform_tags != '' and freeform_tags[-1] == ','):
                freeform_tags = freeform_tags [:-1]
            values_for_column[header].append(freeform_tags)

        return values_for_column

    #Export extra fields
    def export_extra_columns(oci_objs, col_header, sheet_dict, values_for_column):
        value = ""
        if (col_header in sheet_dict.keys()):
            # Check if property exists for any object on that sheet
            for oci_obj in oci_objs:
                try:
                    value = oci_obj.__getattribute__(sheet_dict[col_header])
                    value = commonTools.check_exported_value(value)
                    break
                except AttributeError as e:
                    pass
            values_for_column[col_header].append(value)
        # For Cols not defined in Excel_Columns sheet
        else:
            # Check if property exists for any object on that sheet
            for oci_obj in oci_objs:
                try:
                    value = oci_obj.__getattribute__(commonTools.check_column_headers(col_header))
                    value = commonTools.check_exported_value(value)
                    break
                except AttributeError as e:
                    pass
            values_for_column[col_header].append(value)

        return values_for_column

    # Check CD3 Column headers
    def check_column_headers(var_name):
        # replace special characters and spaces with '_' and convert to lowercase
        # replaces multiple occurrence of '_' to just 1
        var_name = var_name.strip()
        var_name = re.sub('[@!#$%^&*<>?/}{~: \n()|]', '_', var_name).lower()
        var_name = re.sub('_+', '_', var_name).lower()
        return var_name

    #Check TF variable Name
    def check_tf_variable(var_name):
        tfname = re.compile('[^a-zA-Z0-9_-]')
        tfnamestart = re.compile('[A-Za-z]')

        var_name = tfname.sub("-", var_name)
        x = tfnamestart.match(var_name)
        # variable name doesnot start with letter; append with c
        if (x == None):
            var_name = "c" + var_name
        return var_name

    # Process ColumnValues
    def check_columnvalue(columnvalue):

        if str(columnvalue).lower() == 'true' or str(columnvalue).lower() == 'false':
            columnvalue = str(columnvalue).lower()

        if (columnvalue.lower() == 'nan'):
            columnvalue = ""

        return columnvalue

    # Process column values with ::
    def check_multivalues_columnvalue(columnvalue, columnname, tempdict):
        columnvalue = str(columnvalue).strip()
        if "::" in columnvalue:
            if ".Flex" in columnvalue:
                columnname = commonTools.check_column_headers(columnname)
                multivalues = columnvalue.split("::")
                multivalues = [str(part).strip() for part in multivalues if part]
                tempdict = {columnname: multivalues}
            elif columnname != 'Compartment Name':
                columnname = commonTools.check_column_headers(columnname)
                multivalues = columnvalue.split("::")
                multivalues = [str(part).strip() for part in multivalues if part]
                tempdict = {columnname: multivalues}
        return tempdict

    # Split values for tagging
    def split_tag_values(columnname, columnvalue, tempdict):
        columnvalue = columnvalue.replace("\n", "")
        # if comma is used as delimeter, replace it with ';'
        if ',' in columnvalue:
            columnvalue = columnvalue.replace(',',';')
        if ";" in columnvalue:
            # If there are more than one tag; split them by ";" and "="
            columnname = commonTools.check_column_headers(columnname)
            multivalues = columnvalue.split(";")
            multivalues = [part.split("=") for part in multivalues if part]
            tempdict = {columnname: multivalues}
        else:
            # If there is only one tag; split them only by "="; each key-value pair is stored as a list
            columnname = commonTools.check_column_headers(columnname)
            multivalues = columnvalue.split("=")
            multivalues = [str(part).strip() for part in multivalues if part]
            tempdict = {columnname: [multivalues]}
        return tempdict

    # Read rows from CD3
    def read_cd3(cd3file, sheet_name):
        if (".xls" not in cd3file or ".xlsx" not in cd3file):
            print("Invalid CD3 Format..Exiting!!!")
            exit()
        df = pd.read_excel(cd3file, sheet_name=sheet_name, skiprows=1, dtype=object)
        yield df
        book = load_workbook(cd3file)
        sheet = book[sheet_name]
        values_for_column = collections.OrderedDict()
        # values_for_column={}
        for j in range(0, sheet.max_column):
            col_name = sheet.cell(row=2, column=j + 1).value
            if (type(col_name) == str):
                values_for_column[col_name] = []
        yield values_for_column

    #Write exported  rows to cd3
    def write_to_cd3(values_for_column, cd3file, sheet_name):
        book = load_workbook(cd3file)
        sheet = book[sheet_name]
        if (sheet_name == "VCN Info"):
            onprem_destinations = ""
            ngw_destinations = ""
            igw_destinations = ""
            for destination in values_for_column["onprem_destinations"]:
                onprem_destinations=destination+","+onprem_destinations
            for destination in values_for_column["ngw_destinations"]:
                ngw_destinations = destination + "," + ngw_destinations
            for destination in values_for_column["igw_destinations"]:
                igw_destinations = destination + "," + igw_destinations

            if (onprem_destinations != "" and onprem_destinations[-1] == ','):
                onprem_destinations = onprem_destinations[:-1]
            if (ngw_destinations != "" and ngw_destinations[-1] == ','):
                ngw_destinations = ngw_destinations[:-1]
            if (igw_destinations != "" and igw_destinations[-1] == ','):
                igw_destinations = igw_destinations[:-1]

            sheet.cell(3,2).value = onprem_destinations
            sheet.cell(4,2).value = ngw_destinations
            sheet.cell(5,2).value = igw_destinations
            # Put n for subnet_name_attach_cidr
            sheet.cell(6, 2).value = 'n'
            book.save(cd3file)
            book.close()
            return


        #rows_len=len(rows)
        rows_len = len(values_for_column["Region"])

        #If no rows exported from OCI, remove the sample data as well
        if(rows_len == 0):
            print("0 rows exported; Nothing to write to CD3 excel; Tab "+sheet_name +" will be empty in CD3 excel!!")
            for i in range(0, sheet.max_row):
                for j in range(0, sheet.max_column):
                    sheet.cell(row=i + 3, column=j + 1).value = ""

            book.save(cd3file)
            book.close()
            return

        sheet_max_rows = sheet.max_row
        if (rows_len > sheet_max_rows):
            large = rows_len
        else:
            large = sheet_max_rows

        #Put Data
        j=0
        for i in range(0,large):
            for col_name in values_for_column.keys():
                if(i>=rows_len):
                    sheet.cell(row=i+3, column=j+1).value = ""
                else:
                    sheet.cell(row=i+3, column=j+1).value = values_for_column[col_name][i]
                sheet.cell(row=i+3, column=j+1).alignment = Alignment(wrap_text=True)
                j=j+1
            j=0


        brdr = Border(left=Side(style='thin'),
                      right=Side(style='thin'),
                      top=Side(style='thin'),
                      bottom=Side(style='thin'),
                      )

        for row in sheet.iter_rows(min_row=3):
            for cell in row:
                cell.border = brdr

        # Add color for exported sec rules and route rules
        if (sheet_name == "RouteRulesinOCI" or sheet_name == "SecRulesinOCI"):
            names = []
            # Add color coding to exported rules
            for row in sheet.iter_rows(min_row=3):
                c = 0
                region = ""
                name = ""
                for cell in row:
                    c = c + 1
                    if (c == 1):
                        region = cell.value
                        continue
                    elif (c == 4):
                        name = cell.value
                        break

                vcn_name = region + "_" + name
                if (vcn_name not in names):
                    names.append(vcn_name)
                    for cellnew in row:
                        if (len(names) % 2 == 0):
                            cellnew.fill = PatternFill(start_color="94AFAF", end_color="94AFAF", fill_type="solid")
                            cellnew.border = brdr
                        else:
                            cellnew.fill = PatternFill(start_color="E5DBBE", end_color="E5DBBE", fill_type="solid")
                            cellnew.border = brdr
                else:
                    for cellnew in row:
                        if (len(names) % 2 == 0):
                            cellnew.fill = PatternFill(start_color="94AFAF", end_color="94AFAF", fill_type="solid")
                            cellnew.border = brdr
                        else:
                            cellnew.fill = PatternFill(start_color="E5DBBE", end_color="E5DBBE", fill_type="solid")
                            cellnew.border = brdr

        book.save(cd3file)
        book.close()

    # def backup_file(src_dir, pattern, overwrite):
    def backup_file(src_dir, resource, pattern):
        x = datetime.datetime.now()
        date = x.strftime("%f").strip()
        dest_dir = src_dir + "/backup_" + resource + "_" + datetime.datetime.now().strftime("%d-%m-%H%M%S").replace('/','-')
        for f in os.listdir(src_dir):
            if f.endswith(pattern):
                print("Backing up existing " + pattern + " to " + dest_dir)
                if not os.path.exists(dest_dir):
                    # print("\nCreating backup dir " + dest_dir + "\n")
                    os.makedirs(dest_dir)

                src = os.path.join(src_dir, f)
                dest = os.path.join(dest_dir, f)
                # print("backing up ....." + src +"   to  "+dest)
                shutil.move(src, dest_dir)
                """if (overwrite == 'yes'):
                    shutil.move(src, dest_dir)
                elif (overwrite == 'no'):
                    shutil.copyfile(src, dest)
                """


class parseVCNs():
    peering_dict = dict()

    vcn_region = {}
    vcn_drgs = {}
    vcn_compartment = {}
    vcn_lpg_names = {}
    vcn_lpg_names1 = {}
    vcn_lpg_names2 = {}
    vcn_lpg_names3 = {}
    hub_vcn_names = []
    spoke_vcn_names = []
    vcn_lpg_rules = {}
    vcn_igws = {}
    vcn_ngws = {}
    vcn_sgws = {}
    vcn_hub_spoke_peer_none = {}
    vcn_compartment = {}
    vcn_names = []

    def __init__(self, filename):
        if (".xls" in filename):
            df_vcn = pd.read_excel(filename, sheet_name='VCNs', skiprows=1)
            df_vcn = df_vcn.dropna(how='all')
            df_vcn = df_vcn.reset_index(drop=True)

            # Create VCN details Dicts and Hub and Spoke VCN Names
            for i in df_vcn.index:
                region = df_vcn['Region'][i]
                if (region in commonTools.endNames):  # or str(region).lower() == 'nan'):
                    break
                vcn_name = df_vcn['VCN Name'][i]
                self.vcn_names.append(vcn_name)

                # Check to see if vcn_name is empty in VCNs Sheet
                # if (str(vcn_name).lower() == 'nan'):
                #    print("ERROR!!! vcn_name/row cannot be left empty in VCNs sheet in CD3..exiting...")
                #    exit(1)
                vcn_name = str(vcn_name).strip()
                if str(df_vcn['Hub/Spoke/Peer/None'][i]).strip().split(":")[0].strip().lower() == 'hub':
                    self.peering_dict[vcn_name] = ""

            for i in df_vcn.index:
                region = df_vcn['Region'][i]
                if (region in commonTools.endNames):  # or str(region).lower()=='nan'):
                    break
                vcn_name = df_vcn['VCN Name'][i]
                self.vcn_names.append(vcn_name)

                # Check to see if vcn_name is empty in VCNs Sheet
                # if (str(vcn_name).lower() == 'nan'):
                #    print("ERROR!!! vcn_name/row cannot be left empty in VCNs sheet in CD3..exiting...")
                #    exit(1)
                vcn_name = str(vcn_name).strip()

                region = str(region).strip().lower()
                self.vcn_region[vcn_name] = region

                self.vcn_lpg_names[vcn_name] = str(df_vcn['LPG Required'][i]).strip().split(",")
                self.vcn_lpg_names[vcn_name] = [x.strip() for x in self.vcn_lpg_names[vcn_name]]

                j = 0
                for lpg in self.vcn_lpg_names[vcn_name]:
                    if lpg == 'y':
                        self.vcn_lpg_names[vcn_name][j] = vcn_name + "_lpg" + str(j)
                        j = j + 1
                self.vcn_lpg_names1[vcn_name] = str(df_vcn['LPG Required'][i]).strip().split(",")
                self.vcn_lpg_names1[vcn_name] = [x.strip() for x in self.vcn_lpg_names1[vcn_name]]

                j = 0
                for lpg in self.vcn_lpg_names1[vcn_name]:
                    if lpg == 'y':
                        self.vcn_lpg_names1[vcn_name][j] = vcn_name + "_lpg" + str(j)
                        j = j + 1

                self.vcn_lpg_names2[vcn_name] = str(df_vcn['LPG Required'][i]).strip().split(",")
                self.vcn_lpg_names2[vcn_name] = [x.strip() for x in self.vcn_lpg_names2[vcn_name]]

                j = 0
                for lpg in self.vcn_lpg_names2[vcn_name]:
                    if lpg == 'y':
                        self.vcn_lpg_names2[vcn_name][j] = vcn_name + "_lpg" + str(j)
                        j = j + 1

                self.vcn_lpg_names3[vcn_name] = str(df_vcn['LPG Required'][i]).strip().split(",")
                self.vcn_lpg_names3[vcn_name] = [x.strip() for x in self.vcn_lpg_names3[vcn_name]]

                j = 0
                for lpg in self.vcn_lpg_names3[vcn_name]:
                    if lpg == 'y':
                        self.vcn_lpg_names3[vcn_name][j] = vcn_name + "_lpg" + str(j)
                        j = j + 1

                self.vcn_drgs[vcn_name] = str(df_vcn['DRG Required'][i]).strip()
                self.vcn_igws[vcn_name] = str(df_vcn['IGW Required'][i]).strip()
                self.vcn_ngws[vcn_name] = str(df_vcn['NGW Required'][i]).strip()
                self.vcn_sgws[vcn_name] = str(df_vcn['SGW Required'][i]).strip()
                self.vcn_hub_spoke_peer_none[vcn_name] = str(df_vcn['Hub/Spoke/Peer/None'][i]).strip().split(":")
                self.vcn_compartment[vcn_name] = str(df_vcn['Compartment Name'][i]).strip()

                self.vcn_lpg_rules.setdefault(vcn_name, '')

                if (self.vcn_hub_spoke_peer_none[vcn_name][0].strip().lower() == 'hub'):
                    self.hub_vcn_names.append(vcn_name)
                    # self.peering_dict[vcn_name]=''

                if (self.vcn_hub_spoke_peer_none[vcn_name][0].strip().lower() == 'spoke'):
                    hub_name = self.vcn_hub_spoke_peer_none[vcn_name][1].strip()
                    self.spoke_vcn_names.append(vcn_name)
                    try:
                        self.peering_dict[hub_name] = self.peering_dict[hub_name] + vcn_name + ","
                    except KeyError:
                        print("ERROR!!! "+hub_name +" not marked as Hub. Verify hub_spoke_peer_none column again..Exiting!")
                        exit(1)

                if (self.vcn_hub_spoke_peer_none[vcn_name][0].strip().lower() == 'peer'):
                    self.peering_dict[vcn_name] = self.vcn_hub_spoke_peer_none[vcn_name][1].strip()

            for k, v in self.peering_dict.items():
                if (v != "" and v[-1] == ','):
                    v = v[:-1]
                    self.peering_dict[k] = v

        if (".csv" in filename):
            config = configparser.RawConfigParser()
            config.optionxform = str
            file_read = config.read(filename)
            sections = config.sections()

            # Get Global Properties from Default Section
            all_regions = config.get('Default', 'regions')
            all_regions = all_regions.split(",")
            all_regions = [x.strip().lower() for x in all_regions]


class parseVCNInfo():
    # all_regions = []
    subnet_name_attach_cidr = ''
    onprem_destinations = []
    ngw_destinations = []
    igw_destinations = []

    def __init__(self, filename):

        df_info = pd.read_excel(filename, sheet_name='VCN Info', skiprows=1)

        # Get Property Values
        values = df_info['Value']

        onprem_destinations = str(values[0]).strip()
        if (onprem_destinations.lower() == 'nan'):
            # print("\ndrg_subnet should not be left empty.. It will create empty route tables")
            self.onprem_destinations.append('')
        else:
            self.onprem_destinations = onprem_destinations.split(",")

        ngw_destinations = str(values[1]).strip()
        if (ngw_destinations.lower() == 'nan'):
            self.ngw_destinations.append('0.0.0.0/0')
        else:
            self.ngw_destinations = ngw_destinations.split(",")

        igw_destinations = str(values[2]).strip()
        if (igw_destinations.lower() == 'nan'):
            self.igw_destinations.append('0.0.0.0/0')
        else:
            self.igw_destinations = igw_destinations.split(",")

        self.subnet_name_attach_cidr = str(values[3]).strip()
        if (self.subnet_name_attach_cidr.lower() == 'nan'):
            self.subnet_name_attach_cidr = 'n'
        else:
            self.subnet_name_attach_cidr = self.subnet_name_attach_cidr.strip().lower()

        """all_regions_excel = str(values[4]).strip()
        if(all_regions_excel.lower()=="nan"):
            print("\nERROR!!! regions field in VCN Info tab cannot be left empty..Exiting!!")
            exit(1)
        all_regions_excel = all_regions_excel.split(",")

        self.all_regions = [x.strip().lower() for x in all_regions_excel]
        """

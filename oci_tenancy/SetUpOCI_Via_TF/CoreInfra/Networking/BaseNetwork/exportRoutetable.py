#!/usr/bin/python3


import argparse
import sys
import oci
from oci.core.virtual_network_client import VirtualNetworkClient
from oci.identity import IdentityClient
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Border
from openpyxl.styles import Side
import os
sys.path.append(os.getcwd()+"/../../..")
from commonTools import *

def convertNullToNothing(input):
    EMPTY_STRING = ""
    if input is None:
        return EMPTY_STRING
    else:
        return str(input)

compartment_ids={}
def get_network_compartment_id(config):#, compartment_name):
    identity = IdentityClient(config)
    comp_list = oci.pagination.list_call_get_all_results(identity.list_compartments,config["tenancy"],compartment_id_in_subtree=True)
    compartment_list = comp_list.data

    #if(compartment_name=='root'):
    for compartment in compartment_list:
        if(compartment.lifecycle_state == 'ACTIVE'):
            compartment_ids[compartment.name]=compartment.id
    return compartment_ids

def get_vcn_id(config,compartment_id,vname):
    vcncient = VirtualNetworkClient(config)
    vcnlist = vcncient.list_vcns(compartment_id=compartment_id,lifecycle_state="AVAILABLE")
    vcn_name = vname.lower()
    for v in vcnlist.data:
        name = v.display_name
        if name.lower() == vcn_name:
            return v.id

def get_vcns(config,compartment_id):
    vcncient = VirtualNetworkClient(config)
    vcnlist = vcncient.list_vcns(compartment_id=compartment_id,lifecycle_state="AVAILABLE")
    return vcnlist

def get_network_entity_name(config,network_identity_id):
    vcn1 = VirtualNetworkClient(config)
    if('internetgateway' in network_identity_id):
        igw=vcn1.get_internet_gateway(network_identity_id)
        network_identity_name = "igw:"+igw.data.display_name
        return  network_identity_name

    elif ('servicegateway' in network_identity_id):
        sgw = vcn1.get_service_gateway(network_identity_id)
        network_identity_name = "sgw:"+sgw.data.display_name
        return network_identity_name


    elif ('natgateway' in network_identity_id):
        ngw = vcn1.get_nat_gateway(network_identity_id)
        network_identity_name = "ngw:"+ngw.data.display_name
        return network_identity_name

    elif ('localpeeringgateway' in network_identity_id):
        lpg = vcn1.get_local_peering_gateway(network_identity_id)
        network_identity_name = "lpg:"+lpg.data.display_name
        return network_identity_name

    elif ('drg' in network_identity_id):
        drg = vcn1.get_drg(network_identity_id)
        network_identity_name = "drg:"+drg.data.display_name
        return network_identity_name

        """
    elif ('privateip' in network_identity_id):
        privateip = vcn1.get_private_ip(network_identity_id)
        network_identity_name = "privateip:"+privateip.data.ip_address
        return network_identity_name
        """
    else:
        return network_identity_id

def print_routetables(routetables,region,vcn_name,comp_name):

    #oname.write("SubnetName, Destination CIDR, Route Destination Object, Destination Type\n")
    global i
    global df
    for routetable in routetables.data:
        print(routetable.display_name)
        rules = routetable.route_rules
        display_name = routetable.display_name

        dn=display_name

        #dn = routetable.display_name
        if(not rules):
            i=i+1
            print(i)
            new_row = pd.DataFrame({'Region':region,'Compartment Name':comp_name, 'VCN Name':vcn_name, 'RouteTableName': dn, 'Destination CIDR': '',
                                    'Route Destination Object': '','Destination Type': ''}, index=[i])
            df = df.append(new_row, ignore_index=True)
        for rule in rules:
            i=i+1
            print(i)
            network_entity_id=rule.network_entity_id
            network_entity_name=get_network_entity_name(config,network_entity_id)


            print(dn + "," + str(rule.destination) + "," + str(network_entity_name)+","+ str(rule.destination_type))
            #oname.write(dn + "," + str(rule.destination) + "," +str(rule.network_entity_id)+","+ str(rule.destination_type)+"\n")
            new_row = pd.DataFrame(
                {'Region': region, 'Compartment Name': comp_name, 'VCN Name': vcn_name, 'RouteTableName': dn,
                 'Destination CIDR': str(rule.destination), 'Route Destination Object': str(network_entity_name),
                 'Destination Type': str(rule.destination_type)}, index=[i])
            df = df.append(new_row, ignore_index=True)

parser = argparse.ArgumentParser(description="Export Route Table on OCI to CD3")
parser.add_argument("cd3file", help="path of CD3 excel file to export rules to")
parser.add_argument("--vcnName", help="VCN from which to export the sec list; Leave blank if need to export from all VCNs", required=False)
parser.add_argument("--networkCompartment", help="Compartment where VCN resides", required=False)
parser.add_argument("--configFileName", help="Config file name" , required=False)



if len(sys.argv) < 2:
    parser.print_help()
    sys.exit(1)

args = parser.parse_args()
cd3file=args.cd3file

if('.xls' not in cd3file):
    print("\nAcceptable cd3 format: .xlsx")
    exit()

vcn_name = args.vcnName

if args.configFileName is not None:
    configFileName = args.configFileName
    config = oci.config.from_file(file_location=configFileName)
else:
    config = oci.config.from_file()


ntk_compartment_ids = get_network_compartment_id(config)

i=0
df=pd.DataFrame()
vcnInfo = parseVCNInfo(cd3file)

if vcn_name is not None:
    ntk_comp_name = args.networkCompartment
    if(ntk_comp_name=='' or ntk_comp_name is None):
        print("\nEnter a valid name for the compartment where VCN resides...")
        exit(1)

    found_region = ''
    for reg in vcnInfo.all_regions:
        print("\nSearching for VCN in region..."+reg)
        config.__setitem__("region", commonTools.region_dict[reg])
        vcn_ocid = get_vcn_id(config,ntk_compartment_ids[ntk_comp_name],vcn_name)

        if(vcn_ocid=='' or vcn_ocid is None):
            print('\nCould not find vcn in compartment '+ntk_comp_name+' in region...'+reg)
        else:
            print("\nFound in Region..."+reg)
            found_region=reg
            break

    config.__setitem__("region", commonTools.region_dict[found_region])
    vcn = VirtualNetworkClient(config)
    region=found_region.capitalize()

    routetables = vcn.list_route_tables(compartment_id=ntk_compartment_ids[ntk_comp_name], vcn_id=vcn_ocid, lifecycle_state='AVAILABLE')
    print_routetables(routetables,region, vcn_name,ntk_comp_name)
else:
    print("\nFetching Route Rules for all VCNs in tenancy...")
    for reg in vcnInfo.all_regions:
        for ntk_compartment_name in ntk_compartment_ids:
            config.__setitem__("region", commonTools.region_dict[reg])
            vcn = VirtualNetworkClient(config)
            region = reg.capitalize()
            vcns = get_vcns(config,ntk_compartment_ids[ntk_compartment_name])

            for v in vcns.data:
                vcn_id = v.id
                vcn_name=v.display_name
                routetables = vcn.list_route_tables(compartment_id=ntk_compartment_ids[ntk_compartment_name], vcn_id=vcn_id, lifecycle_state='AVAILABLE')
                print_routetables(routetables,region,vcn_name,ntk_compartment_name)


book = load_workbook(cd3file)
if('RouteRulesinOCI' in book.sheetnames):
    book.remove(book['RouteRulesinOCI'])
writer = pd.ExcelWriter(cd3file, engine='openpyxl')
writer.book = book
writer.save()


book = load_workbook(cd3file)
writer = pd.ExcelWriter(cd3file, engine='openpyxl')
writer.book = book
df.to_excel(writer, sheet_name='RouteRulesinOCI', index=False)

#Adjust column widths
ws=writer.sheets['RouteRulesinOCI']
from openpyxl.utils import get_column_letter
"""dims = {}
for row in ws.rows:
    for cell in row:
        if cell.value:
            dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
for col, value in dims.items():
    ws.column_dimensions[get_column_letter(col)].width = value+1
"""

column_widths = []
for row in ws.rows:
    for i, cell in enumerate(row):
        if len(column_widths) > i:
            if len(str(cell.value)) > column_widths[i]:
                column_widths[i] = len(str(cell.value))
        else:
            column_widths += [len(str(cell.value))]

for i, column_width in enumerate(column_widths):
    ws.column_dimensions[get_column_letter(i+1)].width = column_width

for row in ws.iter_rows(min_row=1, max_row=1, min_col=1):
    for cell in row:
      cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type = "solid")
      cell.font = Font(bold=True)

names=[]
brdr=Border(left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
for row in ws.iter_rows(min_row=2):
    c=0
    region=""
    name=""
    for cell in row:
        c=c+1
        if(c==1):
            region=cell.value
            continue
        elif(c==4):
            name= cell.value
            break

    vcn_name=region+"_"+name
    if(vcn_name not in names):
        names.append(vcn_name)
        for cellnew in row:
            if(len(names) % 2==0):
                cellnew.fill=PatternFill(start_color="94AFAF", end_color="94AFAF", fill_type = "solid")
                cellnew.border=brdr
            else:
                cellnew.fill = PatternFill(start_color="E5DBBE", end_color="E5DBBE", fill_type="solid")
                cellnew.border=brdr
    else:
        for cellnew in row:
            if (len(names) % 2 == 0):
                cellnew.fill = PatternFill(start_color="94AFAF", end_color="94AFAF", fill_type="solid")
                cellnew.border=brdr
            else:
                cellnew.fill = PatternFill(start_color="E5DBBE", end_color="E5DBBE", fill_type="solid")
                cellnew.border=brdr


#Move the sheet near Networking sheets
book._sheets.remove(ws)
book._sheets.insert(9,ws)

writer.save()

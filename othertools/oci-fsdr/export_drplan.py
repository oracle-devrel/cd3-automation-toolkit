import oci
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Alignment, PatternFill, Font
import argparse
from commonLib import *

parser = argparse.ArgumentParser()
parser.add_argument("-o", "--ocid", help="Provide the DR Plan OCID")
parser.add_argument("-s", "--sheet", help="Provide the sheet name under which the value is stored")
parser.add_argument("-f", "--file", help="Provide name of the file to be created/updated")
parser.add_argument("-c", "--config", help="API_KEY")
parser.add_argument("-i", "--instance_principal", help="INSTANCE_PRINCIPAL", nargs='?', const=1, type=int)
parser.add_argument("-t", "--session_token", help="SESSION_TOKEN", nargs='?', const=1, type=int)
args = parser.parse_args()

try:
    #region_file = os.path.dirname(os.path.abspath(__file__))+"/region_file.json"
    #region_map = load_region_map(region_file)
    #region = get_region_from_ocid(args.ocid, region_map)
    region = get_region_from_ocid(args.ocid)
except Exception as e:
    print(f"Error loading region map: {str(e)}")
    exit(1)

try:
    config = oci.config.from_file(file_location=args.config)
except Exception as e:
    print(f"Error loading OCI config: {str(e)}")
    print(".....Exiting!!!")
    exit(1)

if args.ocid:
    config['region'] = region

try:
    if args.instance_principal == 1:
        signer = oci.auth.signers.InstancePrincipalsSecurityTokenSigner()
    elif args.session_token:
        token_file = config['security_token_file']
        token = None
        with open(token_file, 'r') as f:
            token = f.read()

        private_key = oci.signer.load_private_key_from_file(config['key_file'])
        signer = oci.auth.signers.SecurityTokenSigner(token, private_key)
    elif args.config != '':
        signer = oci.signer.Signer(config['tenancy'], config['user'], config['fingerprint'], config['key_file'])
except Exception as e:
    print(f"Error creating signer: {str(e)}")
    exit(1)

try:
    # Get DR Plan
    disaster_recovery_client = oci.disaster_recovery.DisasterRecoveryClient(
        config=config, retry_strategy=oci.retry.DEFAULT_RETRY_STRATEGY, signer=signer)
    get_dr_plan_response = disaster_recovery_client.get_dr_plan(dr_plan_id=args.ocid)
    plan_groups = get_dr_plan_response.data.plan_groups
    # Extract the order of plan groups
    original_order = [pg.id for pg in plan_groups]

    # Manually convert DrPlanGroup objects to dictionaries
    plan_dicts = []
    for pg in plan_groups:
        steps = []
        for step in pg.steps:
            step_dict = {
                'display_name': step.display_name,
                'error_mode': step.error_mode,
                'id': step.id,
                'is_enabled': step.is_enabled,
                'timeout': step.timeout,
                'type': step.type,
            }
            if hasattr(step, 'user_defined_step') and step.user_defined_step:
                user_defined_step = {
                    'step_type': step.user_defined_step.step_type,
                    'run_as_user': getattr(step.user_defined_step, 'run_as_user', None),
                    'run_on_instance_id': getattr(step.user_defined_step, 'run_on_instance_id', None),
                    'function_id': getattr(step.user_defined_step, 'function_id', None),
                    'function_region': getattr(step.user_defined_step, 'function_region', None),
                    'request_body': getattr(step.user_defined_step, 'request_body', None),
                    'object_storage_script_location': {
                        'bucket': getattr(step.user_defined_step.object_storage_script_location, 'bucket', None),
                        'namespace': getattr(step.user_defined_step.object_storage_script_location, 'namespace', None),
                        'object': getattr(step.user_defined_step.object_storage_script_location, 'object', None)
                    } if getattr(step.user_defined_step, 'object_storage_script_location', None) else None,
                    'run_on_instance_region': getattr(step.user_defined_step, 'run_on_instance_region', None),
                    'script_command': getattr(step.user_defined_step, 'script_command', None)
                }
                step_dict['user_defined_step'] = user_defined_step
            steps.append(step_dict)
        plan_dicts.append({
            'display_name': pg.display_name,
            'id': pg.id,
            'type': pg.type,
            'steps': steps
        })

    # Convert the parsed plan data to a DataFrame
    df = pd.json_normalize(plan_dicts)

    # Split the data into two parts based on the "type" value
    built_in_df = df[df['type'] == 'BUILT_IN']
    other_df = df[df['type'] != 'BUILT_IN']

    # Function to normalize and reformat data
    def normalize_and_reformat(df):
        dict_list_orient = df.to_dict('records')
        normalized_data = pd.json_normalize(dict_list_orient, "steps", ['display_name', 'id', 'type'], record_prefix='steps.')
        columns_order = [
            'display_name', 'id', 'steps.display_name', 'steps.error_mode', 'steps.id', 'steps.is_enabled',
            'steps.timeout', 'steps.type', 'type'
        ]
        normalized_data = normalized_data.reindex(columns=columns_order, fill_value=None)
        return normalized_data

    def normalize_other_data(df):
        dict_list_orient = df.to_dict('records')
        normalized_data = pd.json_normalize(dict_list_orient, "steps", ['display_name', 'id', 'type'], record_prefix='steps.')
        columns_order = [
            'display_name', 'id', 'steps.display_name', 'steps.error_mode', 'steps.id', 'steps.is_enabled',
            'steps.timeout', 'steps.type', 'steps.user_defined_step.step_type',
            'steps.user_defined_step.run_as_user', 'steps.user_defined_step.run_on_instance_id',
            'steps.user_defined_step.function_id', 'steps.user_defined_step.function_region', 'steps.user_defined_step.request_body',
            'steps.user_defined_step.object_storage_script_location.bucket', 'steps.user_defined_step.object_storage_script_location.namespace', 'steps.user_defined_step.object_storage_script_location.object',
            'steps.user_defined_step.run_on_instance_region', 'steps.user_defined_step.script_command', 'type'
        ]
        normalized_data = normalized_data.reindex(columns=columns_order, fill_value=None)
        return normalized_data

    # Normalize and reformat both subsets of data
    built_in_data = normalize_and_reformat(built_in_df)
    other_data = normalize_other_data(other_df)

    # Append both subsets of data into one DataFrame
    combined_data = pd.concat([other_data, built_in_data], ignore_index=True)

    # Sort the combined data based on the original order
    combined_data['sort_order'] = pd.Categorical(combined_data['id'], categories=original_order, ordered=True)
    combined_data.sort_values('sort_order', inplace=True)
    combined_data.drop(columns=['sort_order'], inplace=True)

    # Write the combined data to an Excel file
    excel_file = args.file
    sheet = args.sheet
    if sheet.startswith('"') and sheet.endswith('"'):
        sheet = sheet[1:-1]

    # Check if the file exists and the sheet exists
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        if sheet in wb.sheetnames:
            with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                print(f"Writing to sheet: {sheet}")
                combined_data.to_excel(writer, sheet_name=sheet, index=False)
                worksheet = writer.sheets[sheet]
        else:
            with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
                print(f"Writing to sheet: {sheet}")
                combined_data.to_excel(writer, sheet_name=sheet, index=False)
                worksheet = writer.sheets[sheet]
    else:
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='w') as writer:
            print(f"Writing to Excel file: {excel_file} and sheet: {sheet}")
            combined_data.to_excel(writer, sheet_name=sheet, index=False)
            worksheet = writer.sheets[sheet]

    wb = load_workbook(excel_file)
    ws = wb[sheet]


    def merge_and_center(ws, col):
        max_row = ws.max_row
        for row in range(2, max_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            start_row = row
            while row <= max_row and ws.cell(row=row, column=col).value == cell_value:
                row += 1
            end_row = row - 1
            if start_row != end_row:
                ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
                merged_cell = ws.cell(row=start_row, column=col)
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')


    columns_to_merge = ['A', 'B']

    for col in columns_to_merge:
        col_index = column_index_from_string(col)
        merge_and_center(ws, col_index)

    # Define fill colors
    fill_blue = PatternFill(start_color="346EC9", end_color="346EC9", fill_type="solid")
    fill_purple = PatternFill(start_color="858491", end_color="858491", fill_type="solid")
    font_white = Font(color="FFFFFF", bold=True)

    header_cells = ws[1]
    for cell in header_cells:
        if cell.column_letter in ['A', 'B', 'T']:
            cell.fill = fill_blue
            cell.font = font_white
        else:
            cell.fill = fill_purple
            cell.font = font_white

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the modified workbook
    wb.save(excel_file)
    print("Excel file updated successfully.")

    if "Readme" not in wb.sheetnames:
        readme_sheet = wb.create_sheet(title="Readme")
        readme_content = """
        Instructions to update columns in Excel sheet

        For New Plan step update the row values as below:
        - id, steps.id - leave these row values empty column empty 
        - Display_name : Display name for Plan Group name (mandatory)
          steps.display_name : Display name for the step (mandatory)
          steps.error_mode : STOP_ON_ERROR/CONTINUE_ON_ERROR (mandatory)
          steps.is_enabled : TRUE/FALSE (mandatory)
          steps.timeout : timeout value in seconds (mandatory)
          type: USER_DEFINED (mandatory)
          steps.user_defined_step.step_type : RUN_LOCAL_SCRIPT/RUN_OBJECTSTORE_SCRIPT/INVOKE_FUNCTION

          Based on the step type from above fill in the row values as mentioned : 
            RUN_LOCAL_SCRIPT: 
                            - steps.user_defined_step.run_as_user, (description: user as which the script needs to run)
                            - steps.user_defined_step.run_on_instance_id,  (description: Instance OCID where the script is located)
                            - steps.user_defined_step.script_command (description: script command which needs to run)
            RUN_OBJECTSTORE_SCRIPT: 
                            - steps.user_defined_step.run_on_instance_id, (description: Instance OCID where the script is located)
                            - steps.user_defined_step.object_storage_script_location.bucket, (description: OCI bucket name)
                            - steps.user_defined_step.object_storage_script_location.namespace, (description: OCI bucket namespace name)
                            - steps.user_defined_step.object_storage_script_location.object, (description: script name)
                            - steps.user_defined_step.run_on_instance_region, (description: Instance region name)
                            - steps.user_defined_step.script_command (description: script command which needs to run)
            INVOKE_FUNCTION: 
                            - steps.user_defined_step.function_id (description: OCI Function OCID which needs to be invoked)
                            - steps.user_defined_step.function_region (description: OCI Function region)
                            - steps.user_defined_step.request_body (description: OCI Function request body)
        """

        # Insert the content into a single cell (A1)
        readme_sheet["A1"] = readme_content.strip()

        # Expand the row height to accommodate the text
        readme_sheet.row_dimensions[1].height = 750  # You can adjust this value

        # Auto-adjust column width to fit the content
        readme_sheet.column_dimensions['A'].width = 150  # You can adjust this value

        # Set text wrapping for the cell
        readme_sheet["A1"].alignment = Alignment(wrap_text=True, vertical='top')
        readme_sheet["A1"].font = Font(size=14, color="FFFFFF", bold=True)  # Set font size to 14 and color to white
        readme_sheet["A1"].fill = PatternFill(start_color="346EC9", end_color="346EC9",
                                              fill_type="solid")  # Set background to blue
        readme_index = wb.sheetnames.index("Readme")
        wb._sheets.insert(0, wb._sheets.pop(readme_index))

        # Save the workbook with the new Readme sheet
        wb.save(excel_file)



    wb.save(excel_file)
    wb.close()

except Exception as e:
    print(f"Error: {str(e)}")

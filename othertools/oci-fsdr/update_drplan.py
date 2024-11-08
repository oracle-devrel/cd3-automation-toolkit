import oci
import openpyxl
import argparse
import os
from commonLib import *

parser = argparse.ArgumentParser()
parser.add_argument("-o", "--ocid", help="Provide the DR Plan OCID")
parser.add_argument("-s", "--sheet", help="Provide the sheet name under which the value is stored")
parser.add_argument("-f", "--file", help="Provide name of the file to be created/updated")
parser.add_argument("-c", "--config", help="API_KEY")
parser.add_argument("-i", "--instance_principal", help="INSTANCE_PRINCIPAL", nargs='?', const=1, type=int)
parser.add_argument("-t", "--session_token", help="SESSION_TOKEN", nargs='?', const=1, type=int)

args = parser.parse_args()

try:
    """region_file = os.path.dirname(os.path.abspath(__file__)) + "/region_file.json"
    region_map = load_region_map(region_file)
    region = get_region_from_ocid(args.ocid, region_map)"""
    region = get_region_from_ocid(args.ocid)
except Exception as e:
    print(f"Error loading region map or determining region from OCID: {str(e)}")
    print(".....Exiting!!!")
    exit(0)

try:
    config = oci.config.from_file(file_location=args.config)
    if args.ocid:
        config['region'] = region
except Exception as e:
    print(f"Error loading OCI config: {str(e)}")
    print(".....Exiting!!!")
    exit(0)

try:
    if args.instance_principal == 1:
        signer = oci.auth.signers.InstancePrincipalsSecurityTokenSigner()
    elif args.session_token:
        token_file = config['security_token_file']
        token = None
        with open(token_file, 'r') as f:
            token = f.read()

        private_key = oci.signer.load_private_key_from_file(config['key_file'])
    elif args.config != '':
        signer = oci.signer.Signer(config['tenancy'], config['user'], config['fingerprint'], config['key_file'])
except Exception as e:
    print(f"Error initializing signer: {str(e)}")
    print(".....Exiting!!!")
    exit(0)

# Initialize Disaster Recovery client
try:
    disaster_recovery_client = oci.disaster_recovery.DisasterRecoveryClient(config)
except Exception as e:
    print(f"Error initializing Disaster Recovery client: {str(e)}")
    print(".....Exiting!!!")
    exit(0)

# Function to get the actual value of a cell, considering merged cells
def get_merged_cell_value(sheet, row, col):
    cell = sheet.cell(row=row, column=col)
    for merged_cell_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_cell_range:
            merged_cell = sheet.cell(row=merged_cell_range.min_row, column=merged_cell_range.min_col)
            return merged_cell.value
    return cell.value

# Define functions for plan management with appropriate error handling
def new_plan(row, plan_groups_dict):
    try:
        plan_group_display_name = str(row[0])
        id = None
        step_display_name = str(row[2])
        step_error_mode = row[3]
        s_id = None
        step_is_enabled = row[5]
        timeout = row[6]
        step_type = row[8]
        run_as_user = row[9]
        run_on_instance_id = row[10]
        function_id = row[11]
        function_region = row[12]
        request_body = row[13]
        bucket = row[14]
        namespace = row[15]
        bucket_object = row[16]
        instance_region = row[17]
        script_command = row[18]

        if step_type in ["RUN_LOCAL_SCRIPT", "RUN_OBJECTSTORE_SCRIPT", "INVOKE_FUNCTION"]:
            type = 'USER_DEFINED'
        else:
            raise ValueError(f"Invalid step_type: {step_type}. Must be one of RUN_LOCAL_SCRIPT, RUN_OBJECTSTORE_SCRIPT, INVOKE_FUNCTION")

        valid_step_types = [
            'RUN_OBJECTSTORE_SCRIPT_PRECHECK',
            'RUN_LOCAL_SCRIPT_PRECHECK',
            'INVOKE_FUNCTION_PRECHECK',
            'RUN_OBJECTSTORE_SCRIPT',
            'RUN_LOCAL_SCRIPT',
            'INVOKE_FUNCTION'
        ]

        if step_type not in valid_step_types:
            raise ValueError(f"Invalid step_type: {step_type}. Must be one of {valid_step_types}")

        if plan_group_display_name in plan_groups_dict:
            plan_group_details = plan_groups_dict[plan_group_display_name]
        else:
            plan_group_details = oci.disaster_recovery.models.UpdateDrPlanGroupDetails(
                display_name=plan_group_display_name,
                id=id,
                type=type,
                steps=[]
            )
            plan_groups_dict[plan_group_display_name] = plan_group_details

        if step_type == "RUN_LOCAL_SCRIPT":
            step_details = oci.disaster_recovery.models.UpdateDrPlanStepDetails(
                display_name=step_display_name,
                error_mode=step_error_mode,
                id=s_id,
                timeout=timeout,
                is_enabled=step_is_enabled,
                user_defined_step=oci.disaster_recovery.models.UpdateRunLocalScriptUserDefinedStepDetails(
                    step_type=step_type,
                    run_on_instance_id=run_on_instance_id,
                    run_as_user=run_as_user,
                    script_command=script_command
                )
            )
        elif step_type == "RUN_OBJECTSTORE_SCRIPT":
            step_details = oci.disaster_recovery.models.UpdateDrPlanStepDetails(
                display_name=step_display_name,
                error_mode=step_error_mode,
                id=s_id,
                timeout=timeout,
                is_enabled=step_is_enabled,
                user_defined_step=oci.disaster_recovery.models.UpdateRunObjectStoreScriptUserDefinedStepDetails(
                    step_type=step_type,
                    run_on_instance_id=run_on_instance_id,
                    object_storage_script_location=oci.disaster_recovery.models.UpdateObjectStorageScriptLocationDetails(
                        bucket=bucket,
                        namespace=namespace,
                        object=bucket_object
                    )
                )
            )
        elif step_type == "INVOKE_FUNCTION":
            step_details = oci.disaster_recovery.models.UpdateDrPlanStepDetails(
                display_name=step_display_name,
                error_mode=step_error_mode,
                id=s_id,
                timeout=timeout,
                is_enabled=step_is_enabled,
                user_defined_step=oci.disaster_recovery.models.UpdateInvokeFunctionUserDefinedStepDetails(
                    step_type=step_type,
                    function_id=function_id,
                    request_body=request_body
                )
            )
        else:
            raise ValueError(f"Invalid step_type: {step_type}. Must be one of RUN_LOCAL_SCRIPT, RUN_OBJECTSTORE_SCRIPT, INVOKE_FUNCTION")

        if step_details not in plan_group_details.steps:
            plan_group_details.steps.append(step_details)

        return plan_groups_dict, plan_group_details
    except Exception as e:
        print(f"Error in new_plan function: {str(e)}")
        exit(0)

def existing_plan(row, plan_groups_dict):
    try:
        plan_group_display_name = str(row[0])
        id = str(row[1])
        step_display_name = str(row[2])
        step_error_mode = row[3]
        s_id = str(row[4])
        step_is_enabled = row[5]
        timeout = row[6]
        step_type = row[8]
        run_as_user = row[9]
        run_on_instance_id = row[10]
        function_id = row[11]
        function_region = row[12]
        request_body = row[13]
        bucket = row[14]
        namespace = row[15]
        bucket_object = row[16]
        instance_region = row[17]
        script_command = row[18]

        if step_type in ["RUN_LOCAL_SCRIPT", "RUN_OBJECTSTORE_SCRIPT", "INVOKE_FUNCTION"]:
            type = 'USER_DEFINED'
        else:
            raise ValueError(f"Invalid step_type: {step_type}. Must be one of RUN_LOCAL_SCRIPT, RUN_OBJECTSTORE_SCRIPT, INVOKE_FUNCTION")

        valid_step_types = [
            'RUN_OBJECTSTORE_SCRIPT_PRECHECK',
            'RUN_LOCAL_SCRIPT_PRECHECK',
            'INVOKE_FUNCTION_PRECHECK',
            'RUN_OBJECTSTORE_SCRIPT',
            'RUN_LOCAL_SCRIPT',
            'INVOKE_FUNCTION'
        ]

        if step_type not in valid_step_types:
            raise ValueError(f"Invalid step_type: {step_type}. Must be one of {valid_step_types}")

        if id in plan_groups_dict:
            plan_group_details = plan_groups_dict[id]
        else:
            plan_group_details = oci.disaster_recovery.models.UpdateDrPlanGroupDetails(
                display_name=plan_group_display_name,
                id=id,
                type=type,
                steps=[]
            )
            plan_groups_dict[id] = plan_group_details

        if step_type == "RUN_LOCAL_SCRIPT":
            step_details = oci.disaster_recovery.models.UpdateDrPlanStepDetails(
                display_name=step_display_name,
                error_mode=step_error_mode,
                id=s_id,
                timeout=timeout,
                is_enabled=step_is_enabled,
                user_defined_step=oci.disaster_recovery.models.UpdateRunLocalScriptUserDefinedStepDetails(
                    step_type=step_type,
                    run_on_instance_id=run_on_instance_id,
                    run_as_user=run_as_user,
                    script_command=script_command
                )
            )
        elif step_type == "RUN_OBJECTSTORE_SCRIPT":
            step_details = oci.disaster_recovery.models.UpdateDrPlanStepDetails(
                display_name=step_display_name,
                error_mode=step_error_mode,
                id=s_id,
                timeout=timeout,
                is_enabled=step_is_enabled,
                user_defined_step=oci.disaster_recovery.models.UpdateRunObjectStoreScriptUserDefinedStepDetails(
                    step_type=step_type,
                    run_on_instance_id=run_on_instance_id,
                    object_storage_script_location=oci.disaster_recovery.models.UpdateObjectStorageScriptLocationDetails(
                        bucket=bucket,
                        namespace=namespace,
                        object=bucket_object
                    )
                )
            )
        elif step_type == "INVOKE_FUNCTION":
            step_details = oci.disaster_recovery.models.UpdateDrPlanStepDetails(
                display_name=step_display_name,
                error_mode=step_error_mode,
                id=s_id,
                timeout=timeout,
                is_enabled=step_is_enabled,
                user_defined_step=oci.disaster_recovery.models.UpdateInvokeFunctionUserDefinedStepDetails(
                    step_type=step_type,
                    function_id=function_id,
                    request_body=request_body
                )
            )
        else:
            raise ValueError(f"Invalid step_type: {step_type}. Must be one of RUN_LOCAL_SCRIPT, RUN_OBJECTSTORE_SCRIPT, INVOKE_FUNCTION")

        if step_details not in plan_group_details.steps:
            plan_group_details.steps.append(step_details)

        return plan_groups_dict, plan_group_details
    except Exception as e:
        print(f"Error in existing_plan function: {str(e)}")
        exit(0)

def builtin_function(row, plan_groups_dict):
    try:
        plan_group_display_name = str(row[0])
        id = str(row[1])
        step_display_name = str(row[2])
        step_error_mode = row[3]
        s_id = row[4]
        step_is_enabled = row[5]
        timeout = row[6]
        type = row[19]

        valid_builtin_types = ['BUILT_IN', 'BUILT_IN_PRECHECK', 'USER_DEFINED']
        if type not in valid_builtin_types:
            raise ValueError(f"Invalid value for `type`: {type}. Must be one of {valid_builtin_types}")

        if id in plan_groups_dict:
            plan_group_details = plan_groups_dict[id]
        else:
            plan_group_details = oci.disaster_recovery.models.UpdateDrPlanGroupDetails(
                display_name=plan_group_display_name,
                id=id,
                type=type,
                steps=[]
            )
            plan_groups_dict[id] = plan_group_details

        step_details = oci.disaster_recovery.models.UpdateDrPlanStepDetails(
            display_name=step_display_name,
            error_mode=step_error_mode,
            id=s_id,
            timeout=timeout,
            is_enabled=step_is_enabled
        )

        if step_details not in plan_group_details.steps:
            plan_group_details.steps.append(step_details)

        return plan_groups_dict, plan_group_details
    except Exception as e:
        print(f"Error in builtin_function function: {str(e)}")
        exit(0)

try:
    workbook = openpyxl.load_workbook(args.file)
    sheet = args.sheet
    if sheet.startswith('"') and sheet.endswith('"'):
        sheet = sheet[1:-1]
    sheet = workbook[sheet]
except Exception as e:
    print(f"Error loading Excel file or sheet: {str(e)}")
    print(".....Exiting!!!")
    exit(0)

plan_groups_dict = {}
ordered_plan_groups = []

try:
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        row_values = [get_merged_cell_value(sheet, row[0].row, col) for col in range(1, sheet.max_column + 1)]
        id_value = str(row_values[1])
        type_value = str(row_values[19])

        row_values = [None if val in ["None", None] else val for val in row_values]

        if type_value == "USER_DEFINED":
            if id_value == "None":
                plan_groups_dict, plan_group_details = new_plan(row_values, plan_groups_dict)
            else:
                plan_groups_dict, plan_group_details = existing_plan(row_values, plan_groups_dict)
        else:
            plan_groups_dict, plan_group_details = builtin_function(row_values, plan_groups_dict)

        ordered_plan_groups.append(plan_group_details)
except Exception as e:
    print(f"Error processing rows in Excel sheet: {str(e)}")
    print(".....Exiting!!!")
    exit(0)

final_plan_groups = list(plan_groups_dict.values())

try:
    update_dr_plan_details = oci.disaster_recovery.models.UpdateDrPlanDetails(plan_groups=final_plan_groups)
    update_dr_plan_response = disaster_recovery_client.update_dr_plan(
        update_dr_plan_details=update_dr_plan_details,
        dr_plan_id=args.ocid
    )
    print("Update to DR Plan " + args.ocid + "is successful")
except Exception as e:
    print(f"Error updating DR plan: {str(e)}")
    print(".....Exiting!!!")
    exit(0)
